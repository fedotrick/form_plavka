import sys
import os
import re
import logging
import shutil
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLineEdit,
    QPushButton, QMessageBox, QLabel, QScrollArea, QFrame,
    QDateEdit, QComboBox, QTableWidget, QTableWidgetItem,
    QHBoxLayout, QDialog, QFileDialog, QGroupBox, QGridLayout,
    QTabWidget, QTextEdit, QHeaderView
)
from PySide6.QtCore import Qt, QDate
from PySide6 import QtGui
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import pandas as pd
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QGraphicsDropShadowEffect
import time

# В начале файла добавить настройку логирования
def get_application_path():
    if getattr(sys, 'frozen', False):
        # Если приложение - exe-файл
        return os.path.dirname(sys.executable)
    else:
        # Если приложение запущено как скрипт
        return os.path.dirname(os.path.abspath(__file__))

SCRIPT_DIR = get_application_path()
LOG_FILENAME = os.path.join(SCRIPT_DIR, 'plavka.log')
logging.basicConfig(
    filename=LOG_FILENAME,
    level=logging.DEBUG,  # Изменяем уровень на DEBUG для более подробного логирования
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Вынести настройки в отдельные константы
EXCEL_FILENAME = os.path.join(SCRIPT_DIR, 'plavka.xlsx')

# Создаем директорию для данных, если её нет
if not os.path.exists(os.path.dirname(EXCEL_FILENAME)):
    try:
        os.makedirs(os.path.dirname(EXCEL_FILENAME))
        logging.info(f"Создана директория для данных: {os.path.dirname(EXCEL_FILENAME)}")
    except Exception as e:
        logging.error(f"Не удалось создать директорию для данных: {e}")

TEMPERATURE_RANGE = (500, 2000)
TIME_FORMAT = "HH:MM"

# Добавляем новые константы
SEARCH_FIELDS = ['id_plavka', 'Учетный_номер', 'Номер_плавки', 'Наименование_отливки']
EXPORT_FORMATS = {
    'CSV': '*.csv',
    'PDF': '*.pdf',
    'Excel': '*.xlsx'
}

# Добавляем новые константы
BACKUP_DIR = 'backups'
MAX_BACKUPS = 5  # Максимальное количество резервных копий

# Значения по умолчанию для временных полей
DEFAULT_MOVEMENT_TIME = "00:50"
DEFAULT_HEATING_TIME = "00:00"
DEFAULT_CASTING_TIME = "00:00"

# Функция для сохранения данных в Excel
def save_to_excel(data):
    """Сохранение данных в Excel"""
    try:
        logging.debug(f"=== Начало сохранения данных ===")
        logging.debug(f"Путь к файлу Excel: {EXCEL_FILENAME}")
        logging.debug(f"Данные для сохранения: {data}")

        # Создаем бэкап перед сохранением
        backup_file = f"{EXCEL_FILENAME}.bak"
        try:
            if os.path.exists(EXCEL_FILENAME):
                shutil.copy2(EXCEL_FILENAME, backup_file)
                logging.info(f"Создан бэкап файла: {backup_file}")
        except Exception as e:
            logging.warning(f"Не удалось создать бэкап: {str(e)}")

        # Проверяем директорию
        directory = os.path.dirname(EXCEL_FILENAME)
        if not os.path.exists(directory):
            logging.info(f"Создание директории: {directory}")
            try:
                os.makedirs(directory)
                logging.info("Директория успешно создана")
            except Exception as e:
                error_msg = f"Не удалось создать директорию: {str(e)}"
                logging.error(error_msg)
                return False

        try:
            if os.path.exists(EXCEL_FILENAME):
                wb = load_workbook(EXCEL_FILENAME)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                
                # Находим максимальный id
                max_id = 0
                for row in ws.iter_rows(min_row=2):
                    try:
                        row_id = int(row[-1].value or 0)
                        max_id = max(max_id, row_id)
                    except (ValueError, TypeError):
                        continue
                next_id = max_id + 1
            else:
                wb = Workbook()
                ws = wb.active
                headers = ['id_plavka', 'Учетный_номер', 'Плавка_дата',  
                          'Номер_плавки', 'Номер_кластера', 'Старший_смены_плавки', 
                          'Первый_участник_смены_плавки', 'Второй_участник_смены_плавки', 
                          'Третий_участник_смены_плавки', 'Четвертый_участник_смены_плавки', 
                          'Наименование_отливки', 'Тип_эксперемента', 
                          'Сектор_A_опоки', 'Сектор_B_опоки', 'Сектор_C_опоки', 'Сектор_D_опоки',
                          'Плавка_время_прогрева_ковша_A', 'Плавка_время_перемещения_A', 
                          'Плавка_время_заливки_A', 'Плавка_температура_заливки_A',
                          'Плавка_время_прогрева_ковша_B', 'Плавка_время_перемещения_B', 
                          'Плавка_время_заливки_B', 'Плавка_температура_заливки_B',
                          'Плавка_время_прогрева_ковша_C', 'Плавка_время_перемещения_C', 
                          'Плавка_время_заливки_C', 'Плавка_температура_заливки_C',
                          'Плавка_время_прогрева_ковша_D', 'Плавка_время_перемещения_D', 
                          'Плавка_время_заливки_D', 'Плавка_температура_заливки_D',
                          'Комментарий', 'Плавка_время_заливки', 'id']
                next_id = 1
                # Записываем заголовки
                ws.append(headers)

            # Добавляем id к данным
            data['id'] = next_id

            # Добавляем новую строку
            new_row = []
            for header in headers:
                value = data.get(header, '')
                new_row.append(value)
            
            ws.append(new_row)

            # Сохраняем с обработкой ошибок доступа
            max_retries = 3
            retry_delay = 1  # секунды
            
            for attempt in range(max_retries):
                try:
                    wb.save(EXCEL_FILENAME)
                    wb.close()
                    logging.info("Данные успешно сохранены в Excel")
                    
                    # Удаляем бэкап после успешного сохранения
                    if os.path.exists(backup_file):
                        os.remove(backup_file)
                        logging.info("Бэкап файл удален после успешного сохранения")
                    
                    return True
                except PermissionError:
                    if attempt < max_retries - 1:
                        logging.warning(f"Попытка {attempt + 1}: Файл занят, ожидание {retry_delay} сек.")
                        time.sleep(retry_delay)
                    else:
                        raise
                except Exception as e:
                    raise

        except Exception as e:
            error_msg = f"Ошибка при работе с Excel файлом: {str(e)}"
            logging.error(error_msg)
            
            # Восстанавливаем из бэкапа при ошибке
            if os.path.exists(backup_file):
                try:
                    shutil.copy2(backup_file, EXCEL_FILENAME)
                    logging.info("Восстановлен бэкап файла после ошибки")
                except Exception as restore_error:
                    logging.error(f"Не удалось восстановить из бэкапа: {str(restore_error)}")
            
            return False

    except Exception as e:
        error_msg = f"Критическая ошибка при сохранении данных: {str(e)}"
        logging.error(error_msg)
        return False

# Основное окно приложения
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Электронный журнал плавки")
        
        # Устанавливаем светлый фон в стиле Nord
        self.setStyleSheet("""
            QWidget {
                background-color: #eceff4;
                color: #2e3440;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QPushButton {
                background-color: #5e81ac;
                color: #ffffff;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-size: 14px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #81a1c1;
            }
            QPushButton:pressed {
                background-color: #4c566a;
            }
            QLineEdit, QDateEdit, QComboBox, QTextEdit {
                background-color: #ffffff;
                color: #2e3440;
                border: 2px solid #d8dee9;
                border-radius: 4px;
                padding: 6px;
                min-width: 150px;
                font-size: 13px;
            }
            QLineEdit:focus, QDateEdit:focus, QComboBox:focus, QTextEdit:focus {
                border: 2px solid #5e81ac;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #4c566a;
                width: 0;
                height: 0;
                margin-right: 5px;
            }
            QGroupBox {
                border: 2px solid #d8dee9;
                border-radius: 6px;
                margin-top: 1em;
                padding: 15px;
                font-size: 14px;
                font-weight: bold;
                background-color: #e5e9f0;
            }
            QGroupBox::title {
                color: #5e81ac;
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QLabel {
                color: #2e3440;
                font-size: 13px;
                min-width: 120px;
            }
            QTextEdit {
                min-height: 80px;
            }
            /* Стили для полей с температурой */
            QLineEdit[temperature="true"] {
                color: #bf616a;
                font-weight: bold;
                background-color: #fff0f0;
            }
            /* Стили для полей со временем */
            QLineEdit[time="true"] {
                color: #2e7d32;
                background-color: #f0fff0;
            }
            /* Стили для заголовков секторов */
            QGroupBox[sector="true"] {
                background-color: #e5e9f0;
            }
            QGroupBox[sector="true"]::title {
                color: #5e81ac;
                font-size: 15px;
            }
            /* Скроллбары */
            QScrollBar:vertical {
                border: none;
                background-color: #e5e9f0;
                width: 10px;
                margin: 0;
            }
            QScrollBar::handle:vertical {
                background-color: #81a1c1;
                border-radius: 5px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #5e81ac;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0;
                background: none;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }
        """)
        
        # Создаем все виджеты
        self.create_widgets()
        
        # Настраиваем связи между полями
        self.setup_sector_connections()
        
        # Генерируем начальный номер плавки
        self.generate_plavka_number()
        
        # Подключаем сигнал изменения даты к генерации номера плавки
        self.Плавка_дата.dateChanged.connect(self.generate_plavka_number)
        
        # Создаем основной layout
        main_layout = QHBoxLayout()  # Используем горизонтальный layout
        
        # Создаем левую колонку
        left_column = QVBoxLayout()
        left_column.setSpacing(10)
        
        # Создаем правую колонку
        right_column = QVBoxLayout()
        right_column.setSpacing(10)
        
        # Создаем группы для логического разделения элементов
        basic_info_group = QGroupBox("Основная информация")
        participants_group = QGroupBox("Участники")
        casting_group = QGroupBox("Параметры отливки")
        time_group = QGroupBox("Временные параметры")
        comment_group = QGroupBox("Комментарий")
        
        # Создаем grid layouts для каждой группы
        basic_grid = QGridLayout()
        basic_grid.setSpacing(10)
        participants_grid = QGridLayout()
        participants_grid.setSpacing(10)
        casting_grid = QGridLayout()
        casting_grid.setSpacing(10)
        time_grid = QGridLayout()
        time_grid.setSpacing(10)
        
        # Основная информация
        basic_grid.addWidget(QLabel("Дата:"), 0, 0)
        basic_grid.addWidget(self.Плавка_дата, 0, 1)
        basic_grid.addWidget(QLabel("Номер плавки:"), 1, 0)
        basic_grid.addWidget(self.Номер_плавки, 1, 1)
        basic_grid.addWidget(QLabel("Номер кластера:"), 2, 0)
        basic_grid.addWidget(self.Номер_кластера, 2, 1)
        basic_grid.addWidget(QLabel("Время слива:"), 3, 0)
        basic_grid.addWidget(self.Плавка_время_заливки, 3, 1)
        basic_info_group.setLayout(basic_grid)
        
        # Участники
        participants_grid.addWidget(QLabel("Старший смены:"), 0, 0)
        participants_grid.addWidget(self.Старший_смены_плавки, 0, 1)
        participants_grid.addWidget(QLabel("Участник 1:"), 1, 0)
        participants_grid.addWidget(self.Первый_участник_смены_плавки, 1, 1)
        participants_grid.addWidget(QLabel("Участник 2:"), 2, 0)
        participants_grid.addWidget(self.Второй_участник_смены_плавки, 2, 1)
        participants_grid.addWidget(QLabel("Участник 3:"), 3, 0)
        participants_grid.addWidget(self.Третий_участник_смены_плавки, 3, 1)
        participants_grid.addWidget(QLabel("Участник 4:"), 4, 0)
        participants_grid.addWidget(self.Четвертый_участник_смены_плавки, 4, 1)
        participants_group.setLayout(participants_grid)
        
        # Параметры отливки
        casting_grid.addWidget(QLabel("Наименование:"), 0, 0)
        casting_grid.addWidget(self.Наименование_отливки, 0, 1, 1, 3)
        casting_grid.addWidget(QLabel("Тип эксперимента:"), 1, 0)
        casting_grid.addWidget(self.Тип_эксперемента, 1, 1, 1, 3)
        
        # Добавляем секторы опок в сетку
        casting_grid.addWidget(QLabel("Секторы опок:"), 2, 0)
        sectors_grid = QGridLayout()
        sectors_grid.addWidget(QLabel("A:"), 0, 0)
        sectors_grid.addWidget(self.Сектор_A_опоки, 0, 1)
        sectors_grid.addWidget(QLabel("B:"), 0, 2)
        sectors_grid.addWidget(self.Сектор_B_опоки, 0, 3)
        sectors_grid.addWidget(QLabel("C:"), 1, 0)
        sectors_grid.addWidget(self.Сектор_C_опоки, 1, 1)
        sectors_grid.addWidget(QLabel("D:"), 1, 2)
        sectors_grid.addWidget(self.Сектор_D_опоки, 1, 3)
        sectors_widget = QWidget()
        sectors_widget.setLayout(sectors_grid)
        casting_grid.addWidget(sectors_widget, 2, 1, 1, 3)
        casting_group.setLayout(casting_grid)
        
        # Временные параметры в сетку 2x2
        time_params_layout = QGridLayout()
        
        # Сектор A
        sector_a_group = QGroupBox("Сектор A")
        sector_a_group.setProperty("sector", "true")
        sector_a_layout = QGridLayout()
        sector_a_layout.addWidget(QLabel("Время прогрева:"), 0, 0)
        sector_a_layout.addWidget(self.Плавка_время_прогрева_ковша_A, 0, 1)
        sector_a_layout.addWidget(QLabel("Время перемещения:"), 1, 0)
        sector_a_layout.addWidget(self.Плавка_время_перемещения_A, 1, 1)
        sector_a_layout.addWidget(QLabel("Время заливки:"), 2, 0)
        sector_a_layout.addWidget(self.Плавка_время_заливки_A, 2, 1)
        sector_a_layout.addWidget(QLabel("Температура:"), 3, 0)
        sector_a_layout.addWidget(self.Плавка_температура_заливки_A, 3, 1)
        sector_a_group.setLayout(sector_a_layout)
        time_params_layout.addWidget(sector_a_group, 0, 0)

        # Сектор B
        sector_b_group = QGroupBox("Сектор B")
        sector_b_group.setProperty("sector", "true")
        sector_b_layout = QGridLayout()
        sector_b_layout.addWidget(QLabel("Время прогрева:"), 0, 0)
        sector_b_layout.addWidget(self.Плавка_время_прогрева_ковша_B, 0, 1)
        sector_b_layout.addWidget(QLabel("Время перемещения:"), 1, 0)
        sector_b_layout.addWidget(self.Плавка_время_перемещения_B, 1, 1)
        sector_b_layout.addWidget(QLabel("Время заливки:"), 2, 0)
        sector_b_layout.addWidget(self.Плавка_время_заливки_B, 2, 1)
        sector_b_layout.addWidget(QLabel("Температура:"), 3, 0)
        sector_b_layout.addWidget(self.Плавка_температура_заливки_B, 3, 1)
        sector_b_group.setLayout(sector_b_layout)
        time_params_layout.addWidget(sector_b_group, 0, 1)

        # Сектор C
        sector_c_group = QGroupBox("Сектор C")
        sector_c_group.setProperty("sector", "true")
        sector_c_layout = QGridLayout()
        sector_c_layout.addWidget(QLabel("Время прогрева:"), 0, 0)
        sector_c_layout.addWidget(self.Плавка_время_прогрева_ковша_C, 0, 1)
        sector_c_layout.addWidget(QLabel("Время перемещения:"), 1, 0)
        sector_c_layout.addWidget(self.Плавка_время_перемещения_C, 1, 1)
        sector_c_layout.addWidget(QLabel("Время заливки:"), 2, 0)
        sector_c_layout.addWidget(self.Плавка_время_заливки_C, 2, 1)
        sector_c_layout.addWidget(QLabel("Температура:"), 3, 0)
        sector_c_layout.addWidget(self.Плавка_температура_заливки_C, 3, 1)
        sector_c_group.setLayout(sector_c_layout)
        time_params_layout.addWidget(sector_c_group, 1, 0)

        # Сектор D
        sector_d_group = QGroupBox("Сектор D")
        sector_d_group.setProperty("sector", "true")
        sector_d_layout = QGridLayout()
        sector_d_layout.addWidget(QLabel("Время прогрева:"), 0, 0)
        sector_d_layout.addWidget(self.Плавка_время_прогрева_ковша_D, 0, 1)
        sector_d_layout.addWidget(QLabel("Время перемещения:"), 1, 0)
        sector_d_layout.addWidget(self.Плавка_время_перемещения_D, 1, 1)
        sector_d_layout.addWidget(QLabel("Время заливки:"), 2, 0)
        sector_d_layout.addWidget(self.Плавка_время_заливки_D, 2, 1)
        sector_d_layout.addWidget(QLabel("Температура:"), 3, 0)
        sector_d_layout.addWidget(self.Плавка_температура_заливки_D, 3, 1)
        sector_d_group.setLayout(sector_d_layout)
        time_params_layout.addWidget(sector_d_group, 1, 1)

        time_group.setLayout(time_params_layout)
        
        # Добавляем поле для комментария
        comment_layout = QVBoxLayout()
        comment_layout.addWidget(self.Комментарий)
        comment_group.setLayout(comment_layout)
        
        # Кнопки управления
        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(self.save_button)
        buttons_layout.addWidget(self.search_button)
        
        # Добавляем группы в колонки
        left_column.addWidget(basic_info_group)
        left_column.addWidget(participants_group)
        left_column.addWidget(casting_group)
        
        right_column.addWidget(time_group)
        right_column.addWidget(comment_group)
        right_column.addLayout(buttons_layout)
        
        # Добавляем колонки в основной layout
        left_widget = QWidget()
        left_widget.setLayout(left_column)
        right_widget = QWidget()
        right_widget.setLayout(right_column)
        
        main_layout.addWidget(left_widget)
        main_layout.addWidget(right_widget)
        
        # Устанавливаем основной layout
        self.setLayout(main_layout)
        
        # Устанавливаем размер окна
        self.setMinimumSize(1600, 850)

    def create_widgets(self):
        """Создание всех виджетов формы"""
        # Создаем основные поля ввода
        self.Плавка_дата = QDateEdit(self)
        self.Плавка_дата.setDisplayFormat("dd.MM.yyyy")
        self.Плавка_дата.setCalendarPopup(True)
        self.Плавка_дата.setDate(QDate.currentDate())
        
        self.Номер_плавки = QLineEdit(self)
        self.Номер_плавки.setReadOnly(True)
        
        self.Номер_кластера = QLineEdit(self)
        
        # Добавляем поле для общего времени заливки
        self.Плавка_время_заливки = QLineEdit(self)
        self.Плавка_время_заливки.setInputMask("99:99")
        self.Плавка_время_заливки.setProperty("time", "true")
        self.Плавка_время_заливки.clear()  # Очищаем значение по умолчанию
        
        # Создаем комбобоксы для участников
        self.Старший_смены_плавки = QComboBox(self)
        self.Старший_смены_плавки.addItem("")  # Пустой элемент по умолчанию
        self.Старший_смены_плавки.addItems([
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев", "отсутствует"
        ])
        
        self.Первый_участник_смены_плавки = QComboBox(self)
        self.Первый_участник_смены_плавки.addItem("")  # Пустой элемент по умолчанию
        self.Первый_участник_смены_плавки.addItems([
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев", "отсутствует"
        ])
        
        self.Второй_участник_смены_плавки = QComboBox(self)
        self.Второй_участник_смены_плавки.addItem("")  # Пустой элемент по умолчанию
        self.Второй_участник_смены_плавки.addItems([
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев", "отсутствует"
        ])
        
        self.Третий_участник_смены_плавки = QComboBox(self)
        self.Третий_участник_смены_плавки.addItem("")  # Пустой элемент по умолчанию
        self.Третий_участник_смены_плавки.addItems([
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев", "отсутствует"
        ])
        
        self.Четвертый_участник_смены_плавки = QComboBox(self)
        self.Четвертый_участник_смены_плавки.addItem("")  # Пустой элемент по умолчанию
        self.Четвертый_участник_смены_плавки.addItems([
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев", "отсутствует"
        ])
        
        # Создаем остальные поля
        self.Наименование_отливки = QComboBox(self)
        self.Наименование_отливки.addItem("")  # Пустой элемент по умолчанию
        self.Наименование_отливки.addItems([
            "Вороток", "Ригель", "Ригель optima", "Блок-картер", "Колесо РИТМ",
            "Накладка резьб", "Блок цилиндров", "Диагональ optima", "Кольцо"
        ])
        
        self.Тип_эксперемента = QComboBox(self)
        self.Тип_эксперемента.addItem("")  # Пустой элемент по умолчанию
        self.Тип_эксперемента.addItems(["Бумага", "Волокно"])
        
        # Создаем поля для секторов опок
        self.Сектор_A_опоки = QLineEdit(self)
        self.Сектор_B_опоки = QLineEdit(self)
        self.Сектор_C_опоки = QLineEdit(self)
        self.Сектор_D_опоки = QLineEdit(self)
        
        # Создаем поля для временных параметров сектора A
        self.Плавка_время_прогрева_ковша_A = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_A.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_A.setProperty("time", "true")
        self.Плавка_время_перемещения_A = QLineEdit(self)
        self.Плавка_время_перемещения_A.setInputMask("99:99")
        self.Плавка_время_перемещения_A.setProperty("time", "true")
        self.Плавка_время_заливки_A = QLineEdit(self)
        self.Плавка_время_заливки_A.setInputMask("99:99")
        self.Плавка_время_заливки_A.setProperty("time", "true")
        self.Плавка_температура_заливки_A = QLineEdit(self)
        self.Плавка_температура_заливки_A.setProperty("temperature", "true")
        self.Плавка_температура_заливки_A.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_A.setPlaceholderText("500-2000")

        # Создаем поля для временных параметров сектора B
        self.Плавка_время_прогрева_ковша_B = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_B.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_B.setProperty("time", "true")
        self.Плавка_время_перемещения_B = QLineEdit(self)
        self.Плавка_время_перемещения_B.setInputMask("99:99")
        self.Плавка_время_перемещения_B.setProperty("time", "true")
        self.Плавка_время_заливки_B = QLineEdit(self)
        self.Плавка_время_заливки_B.setInputMask("99:99")
        self.Плавка_время_заливки_B.setProperty("time", "true")
        self.Плавка_температура_заливки_B = QLineEdit(self)
        self.Плавка_температура_заливки_B.setProperty("temperature", "true")
        self.Плавка_температура_заливки_B.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_B.setPlaceholderText("500-2000")

        # Создаем поля для временных параметров сектора C
        self.Плавка_время_прогрева_ковша_C = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_C.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_C.setProperty("time", "true")
        self.Плавка_время_перемещения_C = QLineEdit(self)
        self.Плавка_время_перемещения_C.setInputMask("99:99")
        self.Плавка_время_перемещения_C.setProperty("time", "true")
        self.Плавка_время_заливки_C = QLineEdit(self)
        self.Плавка_время_заливки_C.setInputMask("99:99")
        self.Плавка_время_заливки_C.setProperty("time", "true")
        self.Плавка_температура_заливки_C = QLineEdit(self)
        self.Плавка_температура_заливки_C.setProperty("temperature", "true")
        self.Плавка_температура_заливки_C.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_C.setPlaceholderText("500-2000")

        # Создаем поля для временных параметров сектора D
        self.Плавка_время_прогрева_ковша_D = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_D.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_D.setProperty("time", "true")
        self.Плавка_время_перемещения_D = QLineEdit(self)
        self.Плавка_время_перемещения_D.setInputMask("99:99")
        self.Плавка_время_перемещения_D.setProperty("time", "true")
        self.Плавка_время_заливки_D = QLineEdit(self)
        self.Плавка_время_заливки_D.setInputMask("99:99")
        self.Плавка_время_заливки_D.setProperty("time", "true")
        self.Плавка_температура_заливки_D = QLineEdit(self)
        self.Плавка_температура_заливки_D.setProperty("temperature", "true")
        self.Плавка_температура_заливки_D.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_D.setPlaceholderText("500-2000")

        # Инициализируем состояние полей секторов
        for sector in ['A', 'B', 'C', 'D']:
            fields = [
                getattr(self, f'Плавка_время_прогрева_ковша_{sector}'),
                getattr(self, f'Плавка_время_перемещения_{sector}'),
                getattr(self, f'Плавка_время_заливки_{sector}'),
                getattr(self, f'Плавка_температура_заливки_{sector}')
            ]
            for field in fields:
                field.setEnabled(False)

        # Создаем поле для комментария
        self.Комментарий = QTextEdit(self)
        self.Комментарий.setPlaceholderText("Введите комментарий...")
        
        # Создаем кнопки
        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_data)
        
        self.search_button = QPushButton("Поиск", self)
        self.search_button.clicked.connect(self.show_search_dialog)
        
        # Добавляем обработчик изменения даты
        self.Плавка_дата.dateChanged.connect(self.generate_plavka_number)
        
        # Генерируем начальный номер плавки
        self.generate_plavka_number()

    def setup_sector_connections(self):
        """Настройка связей между полями секторов"""
        # Связываем поля секторов с их обработчиками
        self.Сектор_A_опоки.textChanged.connect(lambda: self.update_sector_fields('A'))
        self.Сектор_B_опоки.textChanged.connect(lambda: self.update_sector_fields('B'))
        self.Сектор_C_опоки.textChanged.connect(lambda: self.update_sector_fields('C'))
        self.Сектор_D_опоки.textChanged.connect(lambda: self.update_sector_fields('D'))
        
        # Связываем температур
        self.Плавка_температура_заливки_A.textChanged.connect(self.sync_temperatures)
        
        # Устанавливаем значения по умолчанию для временных полей
        self.set_default_times()
        
        # Инициализируем начальное состояние полей
        for sector in ['A', 'B', 'C', 'D']:
            # Деактивируем все поля сектора
            fields = [
                getattr(self, f'Плавка_время_прогрева_ковша_{sector}'),
                getattr(self, f'Плавка_время_перемещения_{sector}'),
                getattr(self, f'Плавка_время_заливки_{sector}'),
                getattr(self, f'Плавка_температура_заливки_{sector}')
            ]
            for field in fields:
                field.setEnabled(False)
                field.clear()
            
            # Обновляем состояние полей на основе текущего значения сектора
            self.update_sector_fields(sector)

    def set_default_times(self):
        """Установка значений по умолчанию для временных полей"""
        # Время перемещения
        self.Плавка_время_перемещения_A.setText(DEFAULT_MOVEMENT_TIME)
        self.Плавка_время_перемещения_B.setText(DEFAULT_MOVEMENT_TIME)
        self.Плавка_время_перемещения_C.setText(DEFAULT_MOVEMENT_TIME)
        self.Плавка_время_перемещения_D.setText(DEFAULT_MOVEMENT_TIME)
        
        # Время прогрева ковша
        self.Плавка_время_прогрева_ковша_A.setText(DEFAULT_HEATING_TIME)
        self.Плавка_время_прогрева_ковша_B.setText(DEFAULT_HEATING_TIME)
        self.Плавка_время_прогрева_ковша_C.setText(DEFAULT_HEATING_TIME)
        self.Плавка_время_прогрева_ковша_D.setText(DEFAULT_HEATING_TIME)
        
        # Время заливки
        self.Плавка_время_заливки_A.setText(DEFAULT_CASTING_TIME)
        self.Плавка_время_заливки_B.setText(DEFAULT_CASTING_TIME)
        self.Плавка_время_заливки_C.setText(DEFAULT_CASTING_TIME)
        self.Плавка_время_заливки_D.setText(DEFAULT_CASTING_TIME)

    def update_sector_fields(self, sector):
        """Обновление состояния полей сектора"""
        sector_field = getattr(self, f'Сектор_{sector}_опоки')
        
        # Получаем все поля сектора
        time_fields = {
            'время_перемещения': (getattr(self, f'Плавка_время_перемещения_{sector}'), DEFAULT_MOVEMENT_TIME),
            'время_прогрева': (getattr(self, f'Плавка_время_прогрева_ковша_{sector}'), DEFAULT_HEATING_TIME),
            'время_заливки': (getattr(self, f'Плавка_время_заливки_{sector}'), DEFAULT_CASTING_TIME)
        }
        temp_field = getattr(self, f'Плавка_температура_заливки_{sector}')
        
        is_active = bool(sector_field.text().strip())
        
        # Активируем/деактивируем поля
        for field_type, (field, default_value) in time_fields.items():
            field.setEnabled(is_active)
            if not is_active:
                field.clear()  # Очищаем значение если сектор неактивен
            else:
                # Устанавливаем значение по умолчанию при активации
                field.setText(default_value)
        
        # Обрабатываем поле температур
        temp_field.setEnabled(is_active)
        if not is_active:
            temp_field.clear()
        elif sector != 'A':
            # Для неглавных секторов копируем температур из A
            temp_a = self.Плавка_температура_заливки_A.text()
            if temp_a:
                temp_field.setText(temp_a)

    def sync_temperatures(self):
        """Синхронизация температур всех активных секторов с сектором A"""
        temp_a = self.Плавка_температура_заливки_A.text()
        if not temp_a:
            return
            
        # Проверяем, что температур в допустимом диапазоне
        try:
            temp_value = int(temp_a)
            if not (500 <= temp_value <= 2000):
                return
        except ValueError:
            return
            
        for sector in ['B', 'C', 'D']:
            sector_field = getattr(self, f'Сектор_{sector}_опоки')
            if sector_field.text().strip():  # Если сектор активен
                temp_field = getattr(self, f'Плавка_температура_заливки_{sector}')
                temp_field.setText(temp_a)

    def generate_plavka_number(self):
        current_month = self.Плавка_дата.date().month()
        current_year = self.Плавка_дата.date().year()
        next_number = 1
        
        try:
            if os.path.exists(EXCEL_FILENAME):
                df = pd.read_excel(EXCEL_FILENAME)
                if not df.empty:
                    # Конвертируем даты в datetime
                    df['Плавка_дата'] = pd.to_datetime(df['Плавка_дата'], format='%d.%m.%Y')
                    
                    # Фильтруем записи только текущего месяца и года
                    current_month_records = df[
                        (df['Плавка_дата'].dt.month == current_month) & 
                        (df['Плавка_дата'].dt.year == current_year)
                    ]
                    
                    if not current_month_records.empty:
                        # Ищем последний номер для текущего месяца
                        last_numbers = []
                        for num in current_month_records['Номер_плавки']:
                            try:
                                if isinstance(num, str) and '-' in num:
                                    month_str, number = num.split('-')
                                    # Приводим месяц к числу, убирая ведущие нули
                                    month_num = int(month_str)
                                    if month_num == current_month:
                                        last_numbers.append(int(number))
                            except (ValueError, TypeError):
                                continue
                        
                        if last_numbers:
                            next_number = max(last_numbers) + 1
            
            # Форматируем номер плавки: месяц-номер(с ведущими нулями)
            new_plavka_number = f"{current_month}-{str(next_number).zfill(3)}"
            self.Номер_плавки.setText(new_plavka_number)
            
            # Обновляем учетный номер после генерации номера плавки
            self.update_uchet_number()
            
        except Exception as e:
            logging.error(f"Ошибка при генерации номера плавки: {str(e)}")
            # Всё равно пытаемся сгенерировать номер плавки для нового месяца
            try:
                new_plavka_number = f"{current_month}-{str(next_number).zfill(3)}"
                self.Номер_плавки.setText(new_plavka_number)
                self.update_uchet_number()
            except Exception as inner_e:
                logging.error(f"Критическая ошибка при генерации номера плавки: {str(inner_e)}")
                self.Номер_плавки.setText("")

    def update_uchet_number(self):
        """Обновляет учетный номер на основе номера плавки"""
        try:
            plavka_number = self.Номер_плавки.text()
            if plavka_number:
                year = str(self.Плавка_дата.date().year())[-2:]  # Последние 2 цифры года
                uchet_number = f"{plavka_number}/{year}"
                return uchet_number
        except Exception as e:
            logging.error(f"Ошибка при обновлении учетного номера: {str(e)}")
        return None

    def id_plavka_generation(self, Плавка_дата, Номер_плавки):
        """Генерирует уникальный ID плавки в формате ГГГГММNNN, где:
        ГГГГ - год
        ММ - месяц (01-12)
        NNN - номер плавки (001-999)
        """
        try:
            # Извлекаем число после дефиса
            match = re.search(r'-(\d+)', Номер_плавки)
            if not match:
                logging.error(f"Неверный формат номера плавки: {Номер_плавки}")
                QMessageBox.warning(self, "Ошибка", 
                    "Неверный формат номера плавки. Требуется формат с дефисом (например: 2-123)")
                return None
                
            number = match.group(1)
            if not number or len(number) > 3:
                logging.error(f"Некорректный номер плавки: {number}")
                QMessageBox.warning(self, "Ошибка", "Номер плавки после дефиса должен быть от 1 до 999")
                return None
                
            # Форматируем составляющие id_plavka
            year = str(Плавка_дата.year())  # Полный год (4 цифры)
            month = f"{Плавка_дата.month():02d}"  # Месяц с ведущим нулем
            number = number.zfill(3)  # Номер с ведущими нулями до 3 знаков
            
            id_plavka = f"{year}{month}{number}"
            logging.debug(f"Сгенерирован id_plavka: {id_plavka}")
            return id_plavka
                
        except Exception as e:
            logging.error(f"Ошибка при генерации ID плавки: {str(e)}")
            QMessageBox.warning(self, "Ошибка", f"Не удалось сгенерировать ID плавки: {str(e)}")
            return None

    def generate_учетный_номер(self, Плавка_дата, Номер_плавки):
        """Генерирует учетный номер в формате 'номер_плавки/гг'
        где гг - последние две цифры года
        """
        try:
            # Получаем последние две цифры года
            last_two_digits_year = str(Плавка_дата.year())[-2:]
            
            # Проверяем, что номер плавки не пустой и имеет правильный формат
            if Номер_плавки and re.match(r'^\d+-\d+$', Номер_плавки):
                return f"{Номер_плавки}/{last_two_digits_year}"
            else:
                logging.error(f"Неверный формат номера плавки для учетного номера: {Номер_плавки}")
                QMessageBox.warning(self, "Ошибка", 
                    "Неверный формат номера плавки. Требуется формат: число-число (например: 2-123)")
        except Exception as e:
            logging.error(f"Ошибка при генерации учетного номера: {str(e)}")
            QMessageBox.warning(self, "Ошибка", f"Не удалось сгенерировать учетный номер: {str(e)}")
        
        return None

    def validate_time(self, time_str):
        """Проверка корректности ввода времени в формате ЧЧ:ММ"""
        if not time_str:  # Пустое значение разрешено
            return True
        
        try:
            # Убираем все пробелы
            time_str = time_str.strip()
            
            # Проверяем формат HH:MM
            if ':' not in time_str:
                return False
            
            hours, minutes = time_str.split(':')
            hours = int(hours)
            minutes = int(minutes)
            
            if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
                return False
                
            # Приводим к формату HH:MM
            return True
        except:
            return False
            
    def format_time(self, time_str):
        """Форматирует время в формат HH:MM"""
        if not time_str:
            return ""
            
        try:
            time_str = time_str.strip()
            if ':' not in time_str:
                return time_str
                
            hours, minutes = time_str.split(':')
            hours = int(hours)
            minutes = int(minutes)
            
            if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
                return time_str
                
            return f"{hours:02d}:{minutes:02d}"
        except:
            return time_str

    def format_temperature(self, temp):
        """Форматирует температур"""
        if not temp:
            return ''
        try:
            return str(int(temp))
        except ValueError:
            return ''

    def save_data(self):
        """Сохранение данных"""
        try:
            if not self.validate_fields():
                return False

            # Собираем данные для сохранения
            current_date = self.Плавка_дата.date()
            current_plavka = self.Номер_плавки.text().strip()
            
            data = {
                'id_plavka': self.id_plavka_generation(current_date, current_plavka),
                'Плавка_дата': current_date.toString("dd.MM.yyyy"),
                'Плавка_время_заливки': self.format_time(self.Плавка_время_заливки.text().strip()),
                'Номер_плавки': current_plavka,
                'Номер_кластера': self.Номер_кластера.text().strip(),
                'Учетный_номер': self.generate_учетный_номер(current_date, current_plavka),
                'Старший_смены_плавки': self.Старший_смены_плавки.currentText(),
                'Первый_участник_смены_плавки': self.Первый_участник_смены_плавки.currentText(),
                'Второй_участник_смены_плавки': self.Второй_участник_смены_плавки.currentText(),
                'Третий_участник_смены_плавки': self.Третий_участник_смены_плавки.currentText(),
                'Четвертый_участник_смены_плавки': self.Четвертый_участник_смены_плавки.currentText(),
                'Наименование_отливки': self.Наименование_отливки.currentText(),
                'Тип_эксперемента': self.Тип_эксперемента.currentText(),
                'Комментарий': self.Комментарий.toPlainText().strip()
            }

            # Обрабатываем данные секторов
            for sector in ['A', 'B', 'C', 'D']:
                sector_field = getattr(self, f'Сектор_{sector}_опоки')
                
                data[f'Сектор_{sector}_опоки'] = sector_field.text().strip()
                
                if sector_field.text().strip():
                    # Если сектор активен, сохраняем все его параметры
                    data[f'Плавка_время_прогрева_ковша_{sector}'] = self.format_time(
                        getattr(self, f'Плавка_время_прогрева_ковша_{sector}').text().strip()
                    )
                    data[f'Плавка_время_перемещения_{sector}'] = self.format_time(
                        getattr(self, f'Плавка_время_перемещения_{sector}').text().strip()
                    )
                    data[f'Плавка_время_заливки_{sector}'] = self.format_time(
                        getattr(self, f'Плавка_время_заливки_{sector}').text().strip()
                    )
                    
                    # Для температур используем значение из сектора A для всех секторов
                    temp = getattr(self, f'Плавка_температура_заливки_{sector}').text().strip()
                    data[f'Плавка_температура_заливки_{sector}'] = self.format_temperature(temp)
                else:
                    # Если сектор неактивен, сохраняем пустые значения
                    data[f'Плавка_время_прогрева_ковша_{sector}'] = ''
                    data[f'Плавка_время_перемещения_{sector}'] = ''
                    data[f'Плавка_время_заливки_{sector}'] = ''
                    data[f'Плавка_температура_заливки_{sector}'] = ''

            # Сохраняем данные в Excel
            if save_to_excel(data):
                QMessageBox.information(self, "Успех", "Данные успешно сохранены")
                
                # Очищаем форму и генерируем новый номер плавки
                self.clear_fields()
                
                # Генерируем новый номер плавки после очистки полей
                self.generate_plavka_number()
                
                return True
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось сохранить данные")
                return False

        except Exception as e:
            logging.error(f"Ошибка при сохранении данных: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при сохранении данных: {str(e)}")
            return False

    def validate_fields(self):
        """Проверка полей на корректность"""
        # Проверяем обязательные поля
        required_fields = [
            (self.Плавка_дата, "Дата"),
            (self.Номер_плавки, "Номер плавки"),
            (self.Номер_кластера, "Номер кластера"),
            (self.Плавка_время_заливки, "Время слива")
        ]
        
        empty_fields = []
        for field, field_name in required_fields:
            if isinstance(field, QDateEdit):
                # Для полей даты просто проверяем, что дата установлена
                continue
            elif isinstance(field, QComboBox):
                # Для комбобоксов проверяем, что выбран элемент
                if not field.currentText().strip():
                    empty_fields.append(field_name)
            else:
                # Для текстовых полей проверяем, что они не пустые
                if not field.text().strip():
                    empty_fields.append(field_name)
        
        if empty_fields:
            QMessageBox.warning(self, "Предупреждение", 
                              "Пожалуйста, заполните следующие обязательные поля:\n- " + 
                              "\n- ".join(empty_fields))
            return False

        # Проверяем, что хотя бы один сектор активен
        active_sectors = []
        for sector in ['A', 'B', 'C', 'D']:
            sector_field = getattr(self, f'Сектор_{sector}_опоки')
            if sector_field.text().strip():  # Если сектор активен
                active_sectors.append(sector)
        
        if not active_sectors:
            QMessageBox.warning(self, "Предупреждение", 
                              "Необходимо заполнить хотя бы один сектор (номер опоки)")
            return False
        
        # Проверяем корректность ввода времени для активных секторов
        for sector in active_sectors:
            time_fields = [
                (getattr(self, f'Плавка_время_прогрева_ковша_{sector}'), f"Время прогрева ковша (сектор {sector})"),
                (getattr(self, f'Плавка_время_перемещения_{sector}'), f"Время перемещения (сектор {sector})"),
                (getattr(self, f'Плавка_время_заливки_{sector}'), f"Время заливки (сектор {sector})")
            ]
            
            for field, field_name in time_fields:
                if not self.validate_time(field.text().strip()):
                    QMessageBox.warning(self, "Предупреждение", 
                                      f"Некорректный формат времени в поле '{field_name}'.\n" +
                                      "Используйте формат ЧЧ:ММ.")
                    return False
                
            # Проверяем температур
            temp_field = getattr(self, f'Плавка_температура_заливки_{sector}')
            if not temp_field.text().strip():
                QMessageBox.warning(self, "Предупреждение", 
                                  f"Не указана температур заливки для сектора {sector}")
                return False
        
        return True

    def clear_fields(self):
        """Очищает все поля формы"""
        try:
            # Сохраняем текущую дату перед очисткой
            current_date = self.Плавка_дата.date()
            
            # Очищаем основные поля
            self.Номер_кластера.clear()
            self.Плавка_время_заливки.clear()
            
            # Сбрасываем комбобоксы участников на первый элемент (пустой)
            self.Старший_смены_плавки.setCurrentIndex(0)
            self.Первый_участник_смены_плавки.setCurrentIndex(0)
            self.Второй_участник_смены_плавки.setCurrentIndex(0)
            self.Третий_участник_смены_плавки.setCurrentIndex(0)
            self.Четвертый_участник_смены_плавки.setCurrentIndex(0)
            
            # Сбрасываем наименование и тип эксперимента на первый элемент (пустой)
            self.Наименование_отливки.setCurrentIndex(0)
            self.Тип_эксперемента.setCurrentIndex(0)
            
            # Очищаем комментарий
            self.Комментарий.clear()
            
            # Очищаем поля секторов
            for sector in ['A', 'B', 'C', 'D']:
                # Очищаем номер опоки
                getattr(self, f'Сектор_{sector}_опоки').clear()
                
                # Очищаем все поля сектора и деактивируем их
                time_fields = [
                    getattr(self, f'Плавка_время_прогрева_ковша_{sector}'),
                    getattr(self, f'Плавка_время_перемещения_{sector}'),
                    getattr(self, f'Плавка_время_заливки_{sector}')
                ]
                for field in time_fields:
                    field.clear()
                    field.setEnabled(False)
                
                # Очищаем и деактивируем поле температур
                temp_field = getattr(self, f'Плавка_температура_заливки_{sector}')
                temp_field.clear()
                temp_field.setEnabled(False)
            
            # Восстанавливаем дату
            self.Плавка_дата.setDate(current_date)
            
            # Устанавливаем значения по умолчанию для временных полей
            self.set_default_times()
            
            logging.info("Форма успешно очищена")
            
        except Exception as e:
            logging.error(f"Ошибка при очистке формы: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при очистке формы: {str(e)}")

    def show_search_dialog(self):
        dialog = SearchDialog(self)
        dialog.exec_()

class StatisticsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Создаем таблицу для отображения данных
        self.data_table = QTableWidget()
        self.data_table.setColumnCount(2)
        self.data_table.setHorizontalHeaderLabels(["Дата", "Температура"])
        layout.addWidget(self.data_table)
        
        # Кнопки для разных типов отображения
        buttons_layout = QHBoxLayout()
        
        temp_button = QPushButton("Температуры")
        temp_button.clicked.connect(lambda: self.show_data('temperature'))
        
        casting_button = QPushButton("Статистика отливок")
        casting_button.clicked.connect(lambda: self.show_data('castings'))
        
        time_button = QPushButton("Временной анализ")
        time_button.clicked.connect(lambda: self.show_data('time'))
        
        buttons_layout.addWidget(temp_button)
        buttons_layout.addWidget(casting_button)
        buttons_layout.addWidget(time_button)
        
        layout.addLayout(buttons_layout)
    
    def show_data(self, data_type):
        self.data_table.setRowCount(0)
        
        try:
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            if data_type == 'temperature':
                self._show_temperature(ws, headers)
            elif data_type == 'castings':
                self._show_castings(ws, headers)
            elif data_type == 'time':
                self._show_time_analysis(ws, headers)
                
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при отображении данных: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при отображении данных: {str(e)}")
    
    def _show_temperature(self, ws, headers):
        self.data_table.setColumnCount(3)
        self.data_table.setHorizontalHeaderLabels(["Дата", "Сектор", "Температура"])
        self.data_table.setSortingEnabled(True)  # Включаем сортировку
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            date = data.get('Плавка_дата', '')
            
            # Проверяем температур
            sectors = ['A', 'B', 'C', 'D']
            for sector in sectors:
                try:
                    temp = float(data.get(f'Плавка_температура_заливки_{sector}', 0))
                    if temp > 0:  # Показываем только если есть температур
                        row_position = self.data_table.rowCount()
                        self.data_table.insertRow(row_position)
                        
                        self.data_table.setItem(row_position, 0, QTableWidgetItem(date))
                        self.data_table.setItem(row_position, 1, QTableWidgetItem(f"Сектор {sector}"))
                        self.data_table.setItem(row_position, 2, QTableWidgetItem(f"{temp}°C"))
                except (ValueError, TypeError):
                    continue
        
        if self.data_table.rowCount() == 0:
            QMessageBox.information(self, "Информация", "Нет данных о температур")
            
        self.data_table.resizeColumnsToContents()

    def _show_castings(self, ws, headers):
        """Показывает статистику по отливкам"""
        self.data_table.setColumnCount(2)
        self.data_table.setHorizontalHeaderLabels(["Наименование отливки", "Количество"])
        self.data_table.setSortingEnabled(True)  # Включаем сортировку
        
        # Словарь для подсчета количества каждого типа отливки
        casting_counts = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            casting_name = data.get('Наименование_отливки', '')
            if casting_name:
                casting_counts[casting_name] = casting_counts.get(casting_name, 0) + 1
        
        if not casting_counts:
            QMessageBox.information(self, "Информация", "Нет данных об отливках")
            return
            
        # Отображаем результаты
        for casting_name, count in sorted(casting_counts.items()):
            row_position = self.data_table.rowCount()
            self.data_table.insertRow(row_position)
            
            self.data_table.setItem(row_position, 0, QTableWidgetItem(casting_name))
            self.data_table.setItem(row_position, 1, QTableWidgetItem(str(count)))
        
        self.data_table.resizeColumnsToContents()

    def _show_time_analysis(self, ws, headers):
        """Показывает временной анализ"""
        self.data_table.setColumnCount(4)
        self.data_table.setHorizontalHeaderLabels(["Дата", "Сектор", "Общее время (мин)", "Детализация"])
        self.data_table.setSortingEnabled(True)  # Включаем сортировку
        
        has_data = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            date = data.get('Плавка_дата', '')
            
            # Анализируем время для каждого сектора
            sectors = ['A', 'B', 'C', 'D']
            for sector in sectors:
                try:
                    # Получаем все временные параметры для сектора
                    heating_time = self._convert_time_to_minutes(data.get(f'Плавка_время_прогрева_ковша_{sector}', '0:00'))
                    moving_time = self._convert_time_to_minutes(data.get(f'Плавка_время_перемещения_{sector}', '0:00'))
                    pouring_time = self._convert_time_to_minutes(data.get(f'Плавка_время_заливки_{sector}', '0:00'))
                    
                    # Считаем общее время
                    total_time = heating_time + moving_time + pouring_time
                    
                    if total_time > 0:  # Показываем только если есть какие-то данные
                        has_data = True
                        row_position = self.data_table.rowCount()
                        self.data_table.insertRow(row_position)
                        
                        details = f"Прогрев: {heating_time}мин, Перемещение: {moving_time}мин, Заливка: {pouring_time}мин"
                        
                        self.data_table.setItem(row_position, 0, QTableWidgetItem(date))
                        self.data_table.setItem(row_position, 1, QTableWidgetItem(f"Сектор {sector}"))
                        self.data_table.setItem(row_position, 2, QTableWidgetItem(str(total_time)))
                        self.data_table.setItem(row_position, 3, QTableWidgetItem(details))
                
                except (ValueError, TypeError):
                    continue
        
        if not has_data:
            QMessageBox.information(self, "Информация", "Нет данных о временных параметрах")
            
        self.data_table.resizeColumnsToContents()

    def _convert_time_to_minutes(self, time_str):
        """Конвертирует время из формата ЧЧ:ММ в минуты"""
        try:
            if not time_str or time_str == '0:00':
                return 0
            hours, minutes = map(int, time_str.split(':'))
            return hours * 60 + minutes
        except (ValueError, TypeError):
            return 0

class SearchDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Поиск записей")
        self.setMinimumSize(1000, 700)
        
        # Добавляем тень для окна
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(0)
        shadow.setColor(QColor(0, 0, 0, 60))
        self.setGraphicsEffect(shadow)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Добавляем фильтры
        filter_group = QGroupBox("Фильтры")
        filter_layout = QGridLayout()
        
        # Фильтр по дате
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd.MM.yyyy")
        self.date_from.dateChanged.connect(self.clear_results)
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd.MM.yyyy")
        self.date_to.setDate(QDate.currentDate())
        self.date_to.dateChanged.connect(self.clear_results)
        
        filter_layout.addWidget(QLabel("Дата с:"), 0, 0)
        filter_layout.addWidget(self.date_from, 0, 1)
        filter_layout.addWidget(QLabel("по:"), 0, 2)
        filter_layout.addWidget(self.date_to, 0, 3)
        
        # Фильтр по типу отливки
        self.filter_casting = QComboBox()
        self.filter_casting.addItems(["Все"] + [
            "Вороток", "Ригель", "Ригель optima", "Блок-картер", "Колесо РИТМ",
            "Накладка резьб", "Блок цилиндров", "Диагональ optima", "Кольцо"
        ])
        self.filter_casting.currentIndexChanged.connect(self.clear_results)
        filter_layout.addWidget(QLabel("Тип отливки:"), 1, 0)
        filter_layout.addWidget(self.filter_casting, 1, 1)
        
        # Фильтр по участнику
        self.filter_participant = QComboBox()
        self.filter_participant.addItem("Все")
        self.load_participants()
        self.filter_participant.currentIndexChanged.connect(self.clear_results)
        filter_layout.addWidget(QLabel("Участник:"), 1, 2)
        filter_layout.addWidget(self.filter_participant, 1, 3)
        
        # Фильтр по температуре
        self.temp_from = QLineEdit()
        self.temp_to = QLineEdit()
        # Добавляем валидацию температур
        temp_validator = QtGui.QIntValidator(500, 2000)
        self.temp_from.setValidator(temp_validator)
        self.temp_to.setValidator(temp_validator)
        self.temp_from.setPlaceholderText("500-2000")
        self.temp_to.setPlaceholderText("500-2000")
        self.temp_from.textChanged.connect(self.clear_results)
        self.temp_to.textChanged.connect(self.clear_results)
        
        filter_layout.addWidget(QLabel("Температура от:"), 2, 0)
        filter_layout.addWidget(self.temp_from, 2, 1)
        filter_layout.addWidget(QLabel("до:"), 2, 2)
        filter_layout.addWidget(self.temp_to, 2, 3)
        
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        # Добавляем вкладки для результатов и статистики
        self.tab_widget = QTabWidget()
        
        # Вкладка результатов поиска
        search_tab = QWidget()
        search_layout = QVBoxLayout(search_tab)
        
        # Существующие виджеты поиска
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите текст для поиска...")
        search_layout.addWidget(self.search_input)
        
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(len(SEARCH_FIELDS))
        self.results_table.setHorizontalHeaderLabels(SEARCH_FIELDS)
        self.results_table.setSortingEnabled(True)  # Включаем сортировку
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)  # Выбор строк целиком
        self.results_table.setSelectionMode(QTableWidget.SingleSelection)  # Только одна строка
        self.results_table.horizontalHeader().setStretchLastSection(True)  # Растягиваем последний столбец
        # Устанавливаем размеры столбцов
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(header.ResizeMode.ResizeToContents)
        search_layout.addWidget(self.results_table)
        
        self.tab_widget.addTab(search_tab, "Результаты поиска")
        
        # Вкладка статистики
        stats_tab = QWidget()
        stats_layout = QVBoxLayout(stats_tab)
        
        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        stats_layout.addWidget(self.stats_text)
        
        self.tab_widget.addTab(stats_tab, "Статистика")
        
        # Добавляем вкладку визуализации
        viz_tab = StatisticsWidget()
        self.tab_widget.addTab(viz_tab, "Визуализация")
        
        layout.addWidget(self.tab_widget)
        
        # Кнопки
        button_layout = QHBoxLayout()
        self.search_button = QPushButton("Поиск")
        self.edit_button = QPushButton("Редактировать")
        self.edit_button.setEnabled(False)  # Disable by default
        self.export_button = QPushButton("Экспорт")
        self.stats_button = QPushButton("Обновить статистику")
        self.backup_button = QPushButton("Создать резервную копию")
        
        button_layout.addWidget(self.search_button)
        button_layout.addWidget(self.edit_button)
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.stats_button)
        button_layout.addWidget(self.backup_button)
        layout.addLayout(button_layout)
        
        # Подключаем обработчики
        self.search_button.clicked.connect(self.search_records)
        self.edit_button.clicked.connect(self.edit_selected)
        self.export_button.clicked.connect(self.export_results)
        self.stats_button.clicked.connect(self.update_statistics)
        self.backup_button.clicked.connect(self.create_backup)
        self.results_table.itemSelectionChanged.connect(self.on_selection_changed)
        
    def on_selection_changed(self):
        """Enable edit button only when a record is selected"""
        self.edit_button.setEnabled(len(self.results_table.selectedItems()) > 0)
        
    def apply_filters(self, row, headers):
        """Применяет фильтры к записи"""
        try:
            data = dict(zip(headers, row))
            
            # Фильтр по дате
            record_date = QDate.fromString(data['Плавка_дата'], "dd.MM.yyyy")
            if not (self.date_from.date() <= record_date <= self.date_to.date()):
                return False
            
            # Фильтр по типу отливки
            if self.filter_casting.currentText() != "Все" and \
               data['Наименование_отливки'] != self.filter_casting.currentText():
                return False
            
            # Фильтр по участнику
            if self.filter_participant.currentText() != "Все":
                participant = self.filter_participant.currentText()
                if participant not in [
                    data['Старший_смены_плавки'],
                    data['Первый_участник_смены_плавки'],
                    data['Второй_участник_смены_плавки'],
                    data['Третий_участник_смены_плавки'],
                    data['Четвертый_участник_смены_плавки']
                ]:
                    return False
            
            # Фильтр по температуре
            if self.temp_from.text() and self.temp_to.text():
                try:
                    temps = [
                        float(data['Плавка_температура_заливки_A']),
                        float(data['Плавка_температура_заливки_B']),
                        float(data['Плавка_температура_заливки_C']),
                        float(data['Плавка_температура_заливки_D'])
                    ]
                    temp_from = float(self.temp_from.text())
                    temp_to = float(self.temp_to.text())
                    
                    # Проверяем, что хотя бы одна температур в диапазоне
                    if not any(temp_from <= temp <= temp_to for temp in temps if temp):
                        return False
                except ValueError:
                    pass
            
            return True
        except Exception as e:
            logging.error(f"Ошибка при применении фильтров: {str(e)}")
            return False

    def load_participants(self):
        """Загружает список участников из файла Excel"""
        try:
            if not os.path.exists(EXCEL_FILENAME):
                return
                
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            participants = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                participants.add(data['Старший_смены_плавки'])
                participants.add(data['Первый_участник_смены_плавки'])
                participants.add(data['Второй_участник_смены_плавки'])
                participants.add(data['Третий_участник_смены_плавки'])
                participants.add(data['Четвертый_участник_смены_плавки'])
            
            self.filter_participant.addItems(sorted(participants))
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при загрузке списка участников: {str(e)}")
            
    def update_statistics(self):
        """Обновляет статистику по данным"""
        try:
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            stats = {
                'total_records': 0,
                'avg_temp': [],
                'casting_types': {},
                'participants': set(),
                'min_temp': float('inf'),
                'max_temp': float('-inf')
            }
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Проверяем фильтры
                if not self.apply_filters(row, headers):
                    continue
                    
                # Ищем совпадения
                data = dict(zip(headers, row))
                stats['total_records'] += 1
                
                # Температура
                try:
                    temp_A = float(data['Плавка_температура_заливки_A'])
                    temp_B = float(data['Плавка_температура_заливки_B'])
                    temp_C = float(data['Плавка_температура_заливки_C'])
                    temp_D = float(data['Плавка_температура_заливки_D'])
                    stats['avg_temp'].append(temp_A)
                    stats['avg_temp'].append(temp_B)
                    stats['avg_temp'].append(temp_C)
                    stats['avg_temp'].append(temp_D)
                    stats['min_temp'] = min(stats['min_temp'], temp_A, temp_B, temp_C, temp_D)
                    stats['max_temp'] = max(stats['max_temp'], temp_A, temp_B, temp_C, temp_D)
                except (ValueError, TypeError):
                    pass
                
                # Типы отливок
                casting = data['Наименование_отливки']
                stats['casting_types'][casting] = stats['casting_types'].get(casting, 0) + 1
                
                # Участники
                stats['participants'].add(data['Старший_смены_плавки'])
                stats['participants'].add(data['Первый_участник_смены_плавки'])
                stats['participants'].add(data['Второй_участник_смены_плавки'])
                stats['participants'].add(data['Третий_участник_смены_плавки'])
                stats['participants'].add(data['Четвертый_участник_смены_плавки'])
            
            # Формируем отчет
            report = [
                "=== Общая статистика ===",
                f"Всего записей: {stats['total_records']}",
                f"Количество участников: {len(stats['participants'])}",
                "",
                "=== Температура заливки ===",
                f"Средняя: {sum(stats['avg_temp'])/len(stats['avg_temp']):.1f}°C" if stats['avg_temp'] else "Нет данных",
                f"Минимальная: {stats['min_temp']}°C" if stats['min_temp'] != float('inf') else "Нет данных",
                f"Максимальная: {stats['max_temp']}°C" if stats['max_temp'] != float('-inf') else "Нет данных",
                "",
                "=== Распределение по типам отливок ===",
            ]
            
            for casting, count in sorted(stats['casting_types'].items()):
                report.append(f"{casting}: {count} ({count/stats['total_records']*100:.1f}%)")
            
            self.stats_text.setText("\n".join(report))
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при обновлении статистики: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обновлении статистики: {str(e)}")

    def search_records(self):
        search_text = self.search_input.text().lower()
        try:
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            self.results_table.setRowCount(0)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Проверяем фильтры
                if not self.apply_filters(row, headers):
                    continue
                    
                # Ищем совпадения
                for cell in row:
                    if cell and str(cell).lower().find(search_text) != -1:
                        row_position = self.results_table.rowCount()
                        self.results_table.insertRow(row_position)
                        
                        for col, field in enumerate(SEARCH_FIELDS):
                            field_index = headers.index(field)
                            self.results_table.setItem(
                                row_position, col, 
                                QTableWidgetItem(str(row[field_index])))
                        break
            
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при поиске: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске: {str(e)}")

    def edit_selected(self):
        current_row = self.results_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите запись для редактирования")
            return
            
        # Получаем id_plavka выбранной записи
        record_id = self.results_table.item(current_row, 0).text()
        
        # Создаем диалог редактирования
        edit_dialog = EditRecordDialog(record_id, self)
        if edit_dialog.exec_() == QDialog.Accepted:
            # Обновляем таблицу после редактирования
            self.search_records()

    def export_results(self):
        try:
            format_str = "Excel files (*.xlsx);;CSV files (*.csv);;PDF files (*.pdf)"
            file_name, selected_format = QFileDialog.getSaveFileName(
                self, "Экспорт данных", "", format_str
            )
            
            if file_name:
                # Создаем DataFrame из данных таблицы
                data = []
                for row in range(self.results_table.rowCount()):
                    row_data = []
                    for col in range(self.results_table.columnCount()):
                        item = self.results_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)
                
                df = pd.DataFrame(data, columns=SEARCH_FIELDS)
                
                if "xlsx" in selected_format:
                    df.to_excel(file_name, index=False)
                elif "csv" in selected_format:
                    df.to_csv(file_name, index=False, encoding='utf-8-sig')  # Поддержка кириллицы
                elif "pdf" in selected_format:
                    try:
                        import pdfkit
                        # Создаем красивый HTML с использованием стилей
                        html = f"""
                        <html>
                        <head>
                            <meta charset="UTF-8">
                            <style>
                                table {{
                                    width: 100%;
                                    border-collapse: collapse;
                                    margin: 20px 0;
                                    font-family: Arial, sans-serif;
                                }}
                                th, td {{
                                    padding: 12px;
                                    text-align: left;
                                    border-bottom: 1px solid #ddd;
                                }}
                                th {{
                                    background-color: #5e81ac;
                                    color: white;
                                }}
                                tr:nth-child(even) {{
                                    background-color: #f9f9f9;
                                }}
                                tr:hover {{
                                    background-color: #f5f5f5;
                                }}
                                h1 {{
                                    color: #2e3440;
                                    font-family: Arial, sans-serif;
                                    text-align: center;
                                    margin: 20px 0;
                                }}
                            </style>
                        </head>
                        <body>
                            <h1>Отчет по плавкам</h1>
                            {df.to_html(index=False)}
                        </body>
                        </html>
                        """
                        
                        # Конвертируем HTML в PDF
                        pdfkit.from_string(html, file_name, options={
                            'encoding': 'UTF-8',
                            'page-size': 'A4',
                            'margin-top': '20mm',
                            'margin-right': '20mm',
                            'margin-bottom': '20mm',
                            'margin-left': '20mm'
                        })
                    except ImportError:
                        QMessageBox.warning(self, "Предупреждение", 
                            "Для экспорта в PDF требуется установить пакет pdfkit и wkhtmltopdf")
                        return
                    except Exception as e:
                        raise Exception(f"Ошибка при создании PDF: {str(e)}")
                
                QMessageBox.information(self, "Успех", "Данные успешно экспортированы")
                
        except Exception as e:
            logging.error(f"Ошибка при экспорте: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте: {str(e)}")

    def create_backup(self):
        try:
            # Создаем директорию для резервных копий если её нет
            if not os.path.exists(BACKUP_DIR):
                os.makedirs(BACKUP_DIR)
            
            # Формируем имя файла резервной копии
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(BACKUP_DIR, f'plavka_backup_{timestamp}.xlsx')
            
            # Копируем файл
            shutil.copy2(EXCEL_FILENAME, backup_file)
            
            # Удаляем старые резервные копии если их больше MAX_BACKUPS
            backups = sorted([os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR)])
            while len(backups) > MAX_BACKUPS:
                os.remove(backups[0])
                backups.pop(0)
            
            QMessageBox.information(self, "Успех", 
                f"Резервная копия создана:\n{backup_file}")
            
        except Exception as e:
            logging.error(f"Ошибка при создании резервной копии: {str(e)}")
            QMessageBox.critical(self, "Ошибка", 
                f"Ошибка при создании резервной копии: {str(e)}")

    def clear_results(self):
        """Очищает результаты поиска при изменении фильтров"""
        self.results_table.setRowCount(0)
        self.stats_text.clear()

class EditRecordDialog(QDialog):
    def __init__(self, record_id, parent=None):
        super().__init__(parent)
        self.record_id = record_id
        self.setWindowTitle(f"Редактирование записи {record_id}")
        self.setup_ui()
        self.load_record_data()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Создаем область прокрутки
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        content_layout = QVBoxLayout(scroll_content)
        
        # Список участников
        participants = [
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев"
        ]
        participants.sort()
        
        # Список наименований отливок
        naimenovanie_otlivok = [
            "Вороток", "Ригель", "Ригель optima", "Блок-картер", "Колесо РИТМ",
            "Накладка резьб", "Блок цилиндров", "Диагональ optima", "Кольцо"
        ]
        naimenovanie_otlivok.sort()
        
        # Список типов эксперементов
        types = ["Бумага", "Волокно"]
        types.sort()
        
        # Создаем поля ввода
        self.Плавка_дата = QDateEdit(self)
        self.Плавка_дата.setDisplayFormat("dd.MM.yyyy")
        self.Плавка_дата.setCalendarPopup(True)
        content_layout.addWidget(QLabel("Дата:"))
        content_layout.addWidget(self.Плавка_дата)

        self.Номер_плавки = QLineEdit(self)
        content_layout.addWidget(QLabel("Номер плавки:"))
        content_layout.addWidget(self.Номер_плавки)

        self.Номер_кластера = QLineEdit(self)
        content_layout.addWidget(QLabel("Номер кластера:"))
        content_layout.addWidget(self.Номер_кластера)

        # Комбобоксы для участников
        self.Старший_смены_плавки = QComboBox(self)
        self.Старший_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Старший смены:"))
        content_layout.addWidget(self.Старший_смены_плавки)

        self.Первый_участник_смены_плавки = QComboBox(self)
        self.Первый_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Первый участник:"))
        content_layout.addWidget(self.Первый_участник_смены_плавки)

        self.Второй_участник_смены_плавки = QComboBox(self)
        self.Второй_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Второй участник:"))
        content_layout.addWidget(self.Второй_участник_смены_плавки)

        self.Третий_участник_смены_плавки = QComboBox(self)
        self.Третий_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Третий участник:"))
        content_layout.addWidget(self.Третий_участник_смены_плавки)

        self.Четвертый_участник_смены_плавки = QComboBox(self)
        self.Четвертый_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Четвертый участник:"))
        content_layout.addWidget(self.Четвертый_участник_смены_плавки)

        self.Наименование_отливки = QComboBox(self)
        self.Наименование_отливки.addItems(naimenovanie_otlivok)
        content_layout.addWidget(QLabel("Наименование отливки:"))
        content_layout.addWidget(self.Наименование_отливки)

        self.Тип_эксперемента = QComboBox(self)
        self.Тип_эксперемента.addItems(types)
        content_layout.addWidget(QLabel("Тип эксперимента:"))
        content_layout.addWidget(self.Тип_эксперемента)

        # Создаем поля для секторов опок
        self.Сектор_A_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор A опоки:"))
        content_layout.addWidget(self.Сектор_A_опоки)

        self.Сектор_B_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор B опоки:"))
        content_layout.addWidget(self.Сектор_B_опоки)

        self.Сектор_C_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор C опоки:"))
        content_layout.addWidget(self.Сектор_C_опоки)

        self.Сектор_D_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор D опоки:"))
        content_layout.addWidget(self.Сектор_D_опоки)

        # Создаем поля для временных параметров сектора A
        self.Плавка_время_прогрева_ковша_A = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_A.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_A.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время прогрева ковша (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_прогрева_ковша_A)

        self.Плавка_время_перемещения_A = QLineEdit(self)
        self.Плавка_время_перемещения_A.setInputMask("99:99")
        self.Плавка_время_перемещения_A.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время перемещения (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_перемещения_A)

        self.Плавка_время_заливки_A = QLineEdit(self)
        self.Плавка_время_заливки_A.setInputMask("99:99")
        self.Плавка_время_заливки_A.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время заливки (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_заливки_A)

        self.Плавка_температура_заливки_A = QLineEdit(self)
        self.Плавка_температура_заливки_A.setProperty("temperature", "true")
        self.Плавка_температура_заливки_A.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_A.setPlaceholderText("500-2000")
        content_layout.addWidget(QLabel("Температура заливки:"))
        content_layout.addWidget(self.Плавка_температура_заливки_A)

        # Создаем поля для временных параметров сектора B
        self.Плавка_время_прогрева_ковша_B = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_B.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_B.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время прогрева ковша (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_прогрева_ковша_B)

        self.Плавка_время_перемещения_B = QLineEdit(self)
        self.Плавка_время_перемещения_B.setInputMask("99:99")
        self.Плавка_время_перемещения_B.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время перемещения (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_перемещения_B)

        self.Плавка_время_заливки_B = QLineEdit(self)
        self.Плавка_время_заливки_B.setInputMask("99:99")
        self.Плавка_время_заливки_B.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время заливки (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_заливки_B)

        self.Плавка_температура_заливки_B = QLineEdit(self)
        self.Плавка_температура_заливки_B.setProperty("temperature", "true")
        self.Плавка_температура_заливки_B.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_B.setPlaceholderText("500-2000")
        content_layout.addWidget(QLabel("Температура заливки:"))
        content_layout.addWidget(self.Плавка_температура_заливки_B)

        # Создаем поля для временных параметров сектора C
        self.Плавка_время_прогрева_ковша_C = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_C.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_C.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время прогрева ковша (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_прогрева_ковша_C)

        self.Плавка_время_перемещения_C = QLineEdit(self)
        self.Плавка_время_перемещения_C.setInputMask("99:99")
        self.Плавка_время_перемещения_C.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время перемещения (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_перемещения_C)

        self.Плавка_время_заливки_C = QLineEdit(self)
        self.Плавка_время_заливки_C.setInputMask("99:99")
        self.Плавка_время_заливки_C.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время заливки (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_заливки_C)

        self.Плавка_температура_заливки_C = QLineEdit(self)
        self.Плавка_температура_заливки_C.setProperty("temperature", "true")
        self.Плавка_температура_заливки_C.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_C.setPlaceholderText("500-2000")
        content_layout.addWidget(QLabel("Температура заливки:"))
        content_layout.addWidget(self.Плавка_температура_заливки_C)

        # Создаем поля для временных параметров сектора D
        self.Плавка_время_прогрева_ковша_D = QLineEdit(self)
        self.Плавка_время_прогрева_ковша_D.setInputMask("99:99")
        self.Плавка_время_прогрева_ковша_D.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время прогрева ковша (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_прогрева_ковша_D)

        self.Плавка_время_перемещения_D = QLineEdit(self)
        self.Плавка_время_перемещения_D.setInputMask("99:99")
        self.Плавка_время_перемещения_D.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время перемещения (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_перемещения_D)

        self.Плавка_время_заливки_D = QLineEdit(self)
        self.Плавка_время_заливки_D.setInputMask("99:99")
        self.Плавка_время_заливки_D.setProperty("time", "true")
        content_layout.addWidget(QLabel("Время заливки (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_заливки_D)

        self.Плавка_температура_заливки_D = QLineEdit(self)
        self.Плавка_температура_заливки_D.setProperty("temperature", "true")
        self.Плавка_температура_заливки_D.setValidator(QtGui.QIntValidator(500, 2000))
        self.Плавка_температура_заливки_D.setPlaceholderText("500-2000")
        content_layout.addWidget(QLabel("Температура заливки:"))
        content_layout.addWidget(self.Плавка_температура_заливки_D)

        # Создаем поле для комментария
        self.Комментарий = QTextEdit(self)
        self.Комментарий.setPlaceholderText("Введите комментарий...")
        content_layout.addWidget(QLabel("Комментарий:"))
        content_layout.addWidget(self.Комментарий)

        # Кнопки
        button_layout = QHBoxLayout()
        save_button = QPushButton("Сохранить изменения")
        cancel_button = QPushButton("Отмена")
        
        save_button.clicked.connect(self.save_changes)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        
        # Устанавливаем виджеты
        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)
        layout.addLayout(button_layout)

    def load_record_data(self):
        try:
            logging.debug(f"Загрузка записи {self.record_id} из файла {EXCEL_FILENAME}")
            
            if not os.path.exists(EXCEL_FILENAME):
                raise FileNotFoundError(f"Файл {EXCEL_FILENAME} не найден")
                
            try:
                with open(EXCEL_FILENAME, 'r+b') as test_file:
                    pass
            except PermissionError:
                raise PermissionError("Файл Excel открыт в другой программе")
                
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            record_found = False
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == self.record_id:
                    # Заполняем поля данными
                    self.fill_fields(row, headers)
                    record_found = True
                    break
            
            wb.close()
            
            if not record_found:
                raise ValueError(f"Запись с ID {self.record_id} не найдена")
                
            logging.debug(f"Запись {self.record_id} успешно загружена")
            
        except Exception as e:
            error_msg = f"Ошибка при загрузке записи: {str(e)}"
            logging.error(error_msg)
            QMessageBox.critical(self, "Ошибка", error_msg)

    def save_changes(self):
        """Сохраняет изменения в Excel файл"""
        try:
            logging.debug(f"Сохранение изменений для записи {self.record_id}")
            
            if not os.path.exists(EXCEL_FILENAME):
                raise FileNotFoundError(f"Файл {EXCEL_FILENAME} не найден")
            
            # Проверяем доступ к файлу
            try:
                with open(EXCEL_FILENAME, 'r+b') as test_file:
                    pass
            except PermissionError:
                raise PermissionError("Файл Excel открыт в другой программе")
            
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active
            
            # Находим строку с нужным id_plavka
            row_index = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if str(row[0]) == self.record_id:
                    row_index = idx + 2
                    break
            
            if row_index:
                logging.debug(f"Найдена строка {row_index} для записи {self.record_id}")
                
                # Обновляем данные в строке
                ws.cell(row=row_index, column=3).value = self.Плавка_дата.date().toString("dd.MM.yyyy")
                ws.cell(row=row_index, column=4).value = self.format_time(self.Плавка_время_заливки_A.text().strip())
                ws.cell(row=row_index, column=5).value = self.Номер_плавки.text().strip()
                ws.cell(row=row_index, column=6).value = self.Номер_кластера.text().strip()
                ws.cell(row=row_index, column=7).value = self.Старший_смены_плавки.currentText()
                ws.cell(row=row_index, column=8).value = self.Первый_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=9).value = self.Второй_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=10).value = self.Третий_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=11).value = self.Четвертый_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=12).value = self.Наименование_отливки.currentText()
                ws.cell(row=row_index, column=13).value = self.Тип_эксперемента.currentText()
                
                # Сектор A
                ws.cell(row=row_index, column=14).value = self.Сектор_A_опоки.text().strip()
                ws.cell(row=row_index, column=18).value = self.format_time(self.Плавка_время_прогрева_ковша_A.text().strip())
                ws.cell(row=row_index, column=19).value = self.format_time(self.Плавка_время_перемещения_A.text().strip())
                ws.cell(row=row_index, column=20).value = self.format_time(self.Плавка_время_заливки_A.text().strip())
                ws.cell(row=row_index, column=21).value = self.Плавка_температура_заливки_A.text().strip()
                
                # Сектор B
                ws.cell(row=row_index, column=15).value = self.Сектор_B_опоки.text().strip()
                ws.cell(row=row_index, column=22).value = self.format_time(self.Плавка_время_прогрева_ковша_B.text().strip())
                ws.cell(row=row_index, column=23).value = self.format_time(self.Плавка_время_перемещения_B.text().strip())
                ws.cell(row=row_index, column=24).value = self.format_time(self.Плавка_время_заливки_B.text().strip())
                ws.cell(row=row_index, column=25).value = self.Плавка_температура_заливки_B.text().strip()
                
                # Сектор C
                ws.cell(row=row_index, column=16).value = self.Сектор_C_опоки.text().strip()
                ws.cell(row=row_index, column=26).value = self.format_time(self.Плавка_время_прогрева_ковша_C.text().strip())
                ws.cell(row=row_index, column=27).value = self.format_time(self.Плавка_время_перемещения_C.text().strip())
                ws.cell(row=row_index, column=28).value = self.format_time(self.Плавка_время_заливки_C.text().strip())
                ws.cell(row=row_index, column=29).value = self.Плавка_температура_заливки_C.text().strip()
                
                # Сектор D
                ws.cell(row=row_index, column=17).value = self.Сектор_D_опоки.text().strip()
                ws.cell(row=row_index, column=30).value = self.format_time(self.Плавка_время_прогрева_ковша_D.text().strip())
                ws.cell(row=row_index, column=31).value = self.format_time(self.Плавка_время_перемещения_D.text().strip())
                ws.cell(row=row_index, column=32).value = self.format_time(self.Плавка_время_заливки_D.text().strip())
                ws.cell(row=row_index, column=33).value = self.Плавка_температура_заливки_D.text().strip()
                
                # Комментарий
                ws.cell(row=row_index, column=34).value = self.Комментарий.toPlainText().strip()
                
                logging.debug("Попытка сохранения изменений")
                try:
                    wb.save(EXCEL_FILENAME)
                    wb.close()
                    logging.info(f"Изменения для записи {self.record_id} успешно сохранены")
                    QMessageBox.information(self, "Успех", "Изменения сохранены")
                    self.accept()
                except PermissionError:
                    raise PermissionError("Не удалось сохранить файл: он открыт в другой программе")
                except Exception as e:
                    raise Exception(f"Ошибка при сохранении файла: {str(e)}")
            else:
                raise ValueError(f"Запись с ID {self.record_id} не найдена")
            
        except Exception as e:
            error_msg = f"Ошибка при сохранении изменений: {str(e)}"
            logging.error(error_msg)
            QMessageBox.critical(self, "Ошибка", error_msg)

    def fill_fields(self, row, headers):
        """Заполняет поля формы данными из записи"""
        try:
            # Создаем словарь с данными
            data = dict(zip(headers, row))
            
            # Заполняем поля
            self.Плавка_дата.setDate(QDate.fromString(str(data['Плавка_дата']), "dd.MM.yyyy"))
            self.Плавка_время_заливки_A.setText(self.format_time(str(data['Плавка_время_заливки'])))
            self.Номер_плавки.setText(str(data['Номер_плавки']))
            self.Номер_кластера.setText(str(data['Номер_кластера']))
            
            # Устанавливаем значения комбобоксов
            self.Старший_смены_плавки.setCurrentText(str(data['Старший_смены_плавки']))
            self.Первый_участник_смены_плавки.setCurrentText(str(data['Первый_участник_смены_плавки']))
            self.Второй_участник_смены_плавки.setCurrentText(str(data['Второй_участник_смены_плавки']))
            self.Третий_участник_смены_плавки.setCurrentText(str(data['Третий_участник_смены_плавки']))
            self.Четвертый_участник_смены_плавки.setCurrentText(str(data['Четвертый_участник_смены_плавки']))
            
            self.Наименование_отливки.setCurrentText(str(data['Наименование_отливки']))
            self.Тип_эксперемента.setCurrentText(str(data['Тип_эксперемента']))
            
            # Заполняем секторы опоки
            self.Сектор_A_опоки.setText(str(data['Сектор_A_опоки']))
            self.Сектор_B_опоки.setText(str(data['Сектор_B_опоки']))
            self.Сектор_C_опоки.setText(str(data['Сектор_C_опоки']))
            self.Сектор_D_опоки.setText(str(data['Сектор_D_опоки']))
            
            # Заполняем время и температур
            self.Плавка_время_прогрева_ковша_A.setText(self.format_time(str(data['Плавка_время_прогрева_ковша_A'])))
            self.Плавка_время_перемещения_A.setText(self.format_time(str(data['Плавка_время_перемещения_A'])))
            self.Плавка_время_заливки_A.setText(self.format_time(str(data['Плавка_время_заливки_A'])))
            self.Плавка_температура_заливки_A.setText(str(data['Плавка_температура_заливки_A']).replace('°C', ''))
            self.Плавка_время_прогрева_ковша_B.setText(self.format_time(str(data['Плавка_время_прогрева_ковша_B'])))
            self.Плавка_время_перемещения_B.setText(self.format_time(str(data['Плавка_время_перемещения_B'])))
            self.Плавка_время_заливки_B.setText(self.format_time(str(data['Плавка_время_заливки_B'])))
            self.Плавка_температура_заливки_B.setText(str(data['Плавка_температура_заливки_B']).replace('°C', ''))
            self.Плавка_время_прогрева_ковша_C.setText(self.format_time(str(data['Плавка_время_прогрева_ковша_C'])))
            self.Плавка_время_перемещения_C.setText(self.format_time(str(data['Плавка_время_перемещения_C'])))
            self.Плавка_время_заливки_C.setText(self.format_time(str(data['Плавка_время_заливки_C'])))
            self.Плавка_температура_заливки_C.setText(str(data['Плавка_температура_заливки_C']).replace('°C', ''))
            self.Плавка_время_прогрева_ковша_D.setText(self.format_time(str(data['Плавка_время_прогрева_ковша_D'])))
            self.Плавка_время_перемещения_D.setText(self.format_time(str(data['Плавка_время_перемещения_D'])))
            self.Плавка_время_заливки_D.setText(self.format_time(str(data['Плавка_время_заливки_D'])))
            self.Плавка_температура_заливки_D.setText(str(data['Плавка_температура_заливки_D']).replace('°C', ''))
            self.Комментарий.setText(str(data['Комментарий']))
            
        except Exception as e:
            logging.error(f"Ошибка при заполнении полей: {str(e)}")
            raise

    def validate_fields(self):
        """Проверка полей на корректность"""
        try:
            # Проверяем обязательные поля
            if not self.Номер_плавки.text().strip():
                raise ValueError("Номер плавки обязателен")
                
            if not self.Номер_кластера.text().strip():
                raise ValueError("Номер кластера обязателен")
                
            # Проверяем корректность времени
            time_fields = [
                (self.Плавка_время_прогрева_ковша_A, "Время прогрева ковша A"),
                (self.Плавка_время_перемещения_A, "Время перемещения A"),
                (self.Плавка_время_заливки_A, "Время заливки A"),
                (self.Плавка_время_прогрева_ковша_B, "Время прогрева ковша B"),
                (self.Плавка_время_перемещения_B, "Время перемещения B"),
                (self.Плавка_время_заливки_B, "Время заливки B"),
                (self.Плавка_время_прогрева_ковша_C, "Время прогрева ковша C"),
                (self.Плавка_время_перемещения_C, "Время перемещения C"),
                (self.Плавка_время_заливки_C, "Время заливки C"),
                (self.Плавка_время_прогрева_ковша_D, "Время прогрева ковша D"),
                (self.Плавка_время_перемещения_D, "Время перемещения D"),
                (self.Плавка_время_заливки_D, "Время заливки D")
            ]
            
            for field, name in time_fields:
                if field.text().strip() and not self.validate_time(field.text().strip()):
                    raise ValueError(f"Некорректный формат времени в поле {name}")
                    
            # Проверяем температур
            temp_fields = [
                (self.Плавка_температура_заливки_A, "Температура заливки A"),
                (self.Плавка_температура_заливки_B, "Температура заливки B"),
                (self.Плавка_температура_заливки_C, "Температура заливки C"),
                (self.Плавка_температура_заливки_D, "Температура заливки D")
            ]
            
            for field, name in temp_fields:
                temp = field.text().strip()
                if temp:
                    try:
                        temp_val = int(temp)
                        if not (500 <= temp_val <= 2000):
                            raise ValueError(f"Температура в поле {name} должна быть от 500 до 2000°C")
                    except ValueError:
                        raise ValueError(f"Некорректное значение температур в поле {name}")
            
            return True
            
        except ValueError as e:
            QMessageBox.warning(self, "Ошибка валидации", str(e))
            return False

def ensure_excel_file_exists():
    """Проверяет существование файла Excel и создает его если его нет"""
    if not os.path.exists(EXCEL_FILENAME):
        try:
            # Создаем директорию если ее нет
            if not os.path.exists(os.path.dirname(EXCEL_FILENAME)):
                os.makedirs(os.path.dirname(EXCEL_FILENAME))
            
            # Создаем файл Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Records"
            
            # Добавляем заголовки
            headers = ['id_plavka', 'Учетный_номер', 'Плавка_дата',  
                      'Номер_плавки', 'Номер_кластера', 'Старший_смены_плавки', 
                      'Первый_участник_смены_плавки', 'Второй_участник_смены_плавки', 
                      'Третий_участник_смены_плавки', 'Четвертый_участник_смены_плавки', 
                      'Наименование_отливки', 'Тип_эксперемента', 
                      'Сектор_A_опоки', 'Сектор_B_опоки', 'Сектор_C_опоки', 'Сектор_D_опоки',
                      'Плавка_время_прогрева_ковша_A', 'Плавка_время_перемещения_A', 
                      'Плавка_время_заливки_A', 'Плавка_температура_заливки_A',
                      'Плавка_время_прогрева_ковша_B', 'Плавка_время_перемещения_B', 
                      'Плавка_время_заливки_B', 'Плавка_температура_заливки_B',
                      'Плавка_время_прогрева_ковша_C', 'Плавка_время_перемещения_C', 
                      'Плавка_время_заливки_C', 'Плавка_температура_заливки_C',
                      'Плавка_время_прогрева_ковша_D', 'Плавка_время_перемещения_D', 
                      'Плавка_время_заливки_D', 'Плавка_температура_заливки_D',
                      'Комментарий', 'Плавка_время_заливки', 'id']
            ws.append(headers)
            
            # Сохраняем файл
            wb.save(EXCEL_FILENAME)
            logging.info(f"Файл {EXCEL_FILENAME} успешно создан")
        except Exception as e:
            logging.error(f"Ошибка при создании файла {EXCEL_FILENAME}: {str(e)}")
            raise

if __name__ == "__main__":
    try:
        # Настраиваем логирование
        logging.basicConfig(
            filename=LOG_FILENAME,
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        
        # Проверяем существование файла Excel
        ensure_excel_file_exists()
        
        # Запускаем приложение
        app = QApplication(sys.argv)
        window = MainWindow()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        logging.critical(f"Критическая ошибка при запуске приложения: {str(e)}")
        print(f"Критическая ошибка: {str(e)}")
        sys.exit(1)
