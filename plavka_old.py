import sys
import os
import re
import logging
import shutil
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLineEdit,
    QPushButton, QMessageBox, QLabel, QScrollArea, QFrame,
    QDateEdit, QComboBox, QTableWidget, QTableWidgetItem,
    QHBoxLayout, QDialog, QFileDialog, QGroupBox, QGridLayout,
    QTabWidget, QTextEdit
)
from PySide6.QtCore import Qt, QDate
from PySide6 import QtGui
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import pandas as pd
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QGraphicsDropShadowEffect

# В начале файла добавить настройку логирования
logging.basicConfig(
    filename='plavka.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Вынести настройки в отдельные константы
EXCEL_FILENAME = 'plavka.xlsx'
TEMPERATURE_RANGE = (500, 2000)
TIME_FORMAT = "HH:mm"

# Добавляем новые константы
SEARCH_FIELDS = ['ID', 'Учетный_номер', 'Номер_плавки', 'Наименование_отливки']
EXPORT_FORMATS = {
    'CSV': '*.csv',
    'PDF': '*.pdf',
    'Excel': '*.xlsx'
}

# Добавляем новые константы
BACKUP_DIR = 'backups'
MAX_BACKUPS = 5  # Максимальное количество резервных копий

# Функция для сохранения данных в Excel
def save_to_excel(ID, Учетный_номер, Плавка_дата, Номер_плавки, Номер_кластера,
                  Старший_смены_плавки, Первый_участник_смены_плавки,
                  Второй_участник_смены_плавки, Третий_участник_смены_плавки,
                  Четвертый_участник_смены_плавки, Наименование_отливки,
                  Тип_эксперемента, Сектор_A_опоки, Сектор_B_опоки,
                  Сектор_C_опоки, Сектор_D_опоки, Плавка_время_прогрева_ковша,
                  Плавка_время_перемещения, Плавка_время_заливки,
                  Плавка_температура_заливки
                  ):
    file_name = 'plavka.xlsx'

    # Проверка, существует ли файл, и создание новой книги, если нет
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Records"
        sheet.append(["ID", "Учетный_номер", "Плавка_дата", "Номер_плавки", "Номер_кластера",
                       "Старший_смены_плавки", "Первый_участник_смены_плавки",
                       "Второй_участник_смены_плавки", "Третий_участник_смены_плавки",
                       "Четвертый_участник_смены_плавки", "Наименование_отливки",
                       "Тип_эксперемента", "Сектор_A_опоки", "Сектор_B_опоки",
                       "Сектор_C_опоки", "Сектор_D_опоки", "Плавка_время_прогрева_ковша",
                       "Плавка_время_перемещения", "Плавка_время_заливки", 
                       "Плавка_температура_заливки"])
    else:
        workbook = load_workbook(file_name)  # Загрузка существующей книги
        sheet = workbook.active  # Получение активного листа

    # Добавление данных в таблицу
    sheet.append([ID, Учетный_номер, Плавка_дата, Номер_плавки, Номер_кластера,
                  Старший_смены_плавки, Первый_участник_смены_плавки,
                  Второй_участник_смены_плавки, Третий_участник_смены_плавки,
                  Четвертый_участник_смены_плавки, Наименование_отливки,
                  Тип_эксперемента, Сектор_A_опоки, Сектор_B_опоки,
                  Сектор_C_опоки, Сектор_D_опоки, Плавка_время_прогрева_ковша,
                  Плавка_время_перемещения, Плавка_время_заливки,
                  Плавка_температура_заливки])

    workbook.save(file_name)

# Основное окно приложения
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Электронный журнал плавки")
        
        # Устанавливаем современный стиль
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f6fa;
                font-family: 'Segoe UI', Arial;
            }
            
            QLabel {
                color: #2f3542;
                font-size: 12px;
                font-weight: bold;
                margin-top: 5px;
            }
            
            QLineEdit, QDateEdit, QComboBox {
                padding: 8px;
                border: 2px solid #dcdde1;
                border-radius: 5px;
                background-color: white;
                color: #2f3542;
                font-size: 12px;
                margin-bottom: 8px;
            }
            
            QLineEdit:focus, QDateEdit:focus, QComboBox:focus {
                border: 2px solid #5352ed;
            }
            
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            
            QComboBox::down-arrow {
                image: url(down_arrow.png);
                width: 12px;
                height: 12px;
            }
            
            QPushButton {
                background-color: #5352ed;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 13px;
                font-weight: bold;
                margin: 5px;
            }
            
            QPushButton:hover {
                background-color: #3742fa;
            }
            
            QPushButton:pressed {
                background-color: #2f3542;
            }
            
            QPushButton#searchButton {
                background-color: #2ed573;
            }
            
            QPushButton#searchButton:hover {
                background-color: #26ae60;
            }
            
            QGroupBox {
                border: 2px solid #dcdde1;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }
            
            QScrollArea {
                border: none;
            }
            
            QTableWidget {
                border: 2px solid #dcdde1;
                border-radius: 5px;
                background-color: white;
                gridline-color: #f1f2f6;
            }
            
            QTableWidget::item {
                padding: 5px;
            }
            
            QTableWidget::item:selected {
                background-color: #5352ed;
                color: white;
            }
            
            QHeaderView::section {
                background-color: #2f3542;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            
            QTabWidget::pane {
                border: 2px solid #dcdde1;
                border-radius: 5px;
                top: -1px;
            }
            
            QTabBar::tab {
                background-color: #f1f2f6;
                color: #2f3542;
                padding: 8px 15px;
                margin-right: 2px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
            }
            
            QTabBar::tab:selected {
                background-color: #5352ed;
                color: white;
            }
            
            QTextEdit {
                border: 2px solid #dcdde1;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
        """)
        
        # Создаем иконку для окна
        self.setWindowIcon(QtGui.QIcon('icon.png'))  # Нужно добавить файл иконки
        
        # Устанавливаем минимальный размер окна
        self.setMinimumSize(800, 600)

        # Создание области прокрутки
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)

        # Создание виджета для размещения в области прокрутки
        scroll_content = QFrame()
        layout = QVBoxLayout(scroll_content)
        layout.setAlignment(Qt.AlignTop)

        title = QLabel("Электронный журнал плавки")
        title.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Список участников
        participants = [
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев", "отсутствует"
        ]

        # Сортировка списка участников по алфавиту
        participants.sort()
        
        # Список наименований отливок
        naimenovanie_otlivok = [
            "Вороток", "Ригель", "Ригель optima", "Блок-картер",
            "Накладка резьб", "Блок цилиндров", "Диагональ optima", "Кольцо"
        ]
        # Сортировка списка наименований отливок по алфавиту
        naimenovanie_otlivok.sort()
        
        # Список типов эксперементов
        types = [
            "Бумага", "Волокно"
        ]
        # Сортировка списка типов эксперементов по алфавиту
        types.sort()
        
        
        self.Плавка_дата = QDateEdit(self)
        self.Плавка_дата.setDisplayFormat("dd.MM.yyyy")
        self.Плавка_дата.setCalendarPopup(True)
        self.Плавка_дата.setDate(QDate.currentDate().addDays(-1))
        self.Плавка_дата.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Плавка_дата)

        self.Номер_плавки = QLineEdit(self)
        self.Номер_плавки.setReadOnly(True)
        self.Номер_плавки.setPlaceholderText("Номер плавки")
        self.Номер_плавки.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Номер_плавки)

        self.Номер_кластера = QLineEdit(self)
        self.Номер_кластера.setPlaceholderText("Номер кластера")
        self.Номер_кластера.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Номер_кластера)

        # Создание метки
        label = QLabel("Старший смены плавки:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Старший_смены_плавки = QComboBox(self)
        self.Старший_смены_плавки.addItems(participants)
        self.Старший_смены_плавки.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Старший_смены_плавки.setStyleSheet("color: black;")
        self.Старший_смены_плавки.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию        
        layout.addWidget(self.Старший_смены_плавки)        
        
        # Создание метки
        label = QLabel("Первый участник смены плавки:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Первый_участник_смены_плавки = QComboBox(self)
        self.Первый_участник_смены_плавки.addItems(participants)
        self.Первый_участник_смены_плавки.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Первый_участник_смены_плавки.setStyleSheet("color: black;")
        self.Первый_участник_смены_плавки.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        layout.addWidget(self.Первый_участник_смены_плавки)

        # Создание метки
        label = QLabel("Второй участник смены плавки:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Второй_участник_смены_плавки = QComboBox(self)
        self.Второй_участник_смены_плавки.addItems(participants)
        self.Второй_участник_смены_плавки.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Второй_участник_смены_плавки.setStyleSheet("color: black;")
        self.Второй_участник_смены_плавки.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        layout.addWidget(self.Второй_участник_смены_плавки)

        # Создание метки
        label = QLabel("Третий участник смены плавки:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Третий_участник_смены_плавки = QComboBox(self)
        self.Третий_участник_смены_плавки.addItems(participants)
        self.Третий_участник_смены_плавки.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Третий_участник_смены_плавки.setStyleSheet("color: black;")
        self.Третий_участник_смены_плавки.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        layout.addWidget(self.Третий_участник_смены_плавки)

        # Создание метки
        label = QLabel("Четвертый участник смены плавки:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Четвертый_участник_смены_плавки = QComboBox(self)
        self.Четвертый_участник_смены_плавки.addItems(participants)
        self.Четвертый_участник_смены_плавки.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Четвертый_участник_смены_плавки.setStyleSheet("color: black;")
        self.Четвертый_участник_смены_плавки.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        layout.addWidget(self.Четвертый_участник_смены_плавки)

        # Создание метки
        label = QLabel("Наименование отливки:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Наименование_отливки = QComboBox(self)
        self.Наименование_отливки.addItems(naimenovanie_otlivok)
        self.Наименование_отливки.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Наименование_отливки.setStyleSheet("color: black;")
        self.Наименование_отливки.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        layout.addWidget(self.Наименование_отливки)

        # Создание метки
        label = QLabel("Тип эксперимента:", self)
        layout.addWidget(label)

        # Создание комбобокса
        self.Тип_эксперемента = QComboBox(self)
        self.Тип_эксперемента.addItems(types)
        self.Тип_эксперемента.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.Тип_эксперемента.setStyleSheet("color: black;")
        self.Тип_эксперемента.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        layout.addWidget(self.Тип_эксперемента)

        self.Сектор_A_опоки = QLineEdit(self)
        self.Сектор_A_опоки.setPlaceholderText("Сектор A опоки")
        self.Сектор_A_опоки.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Сектор_A_опоки)

        self.Сектор_B_опоки = QLineEdit(self)
        self.Сектор_B_опоки.setPlaceholderText("Сектор B опоки")
        self.Сектор_B_опоки.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Сектор_B_опоки)

        self.Сектор_C_опоки = QLineEdit(self)
        self.Сектор_C_опоки.setPlaceholderText("Сектор C опоки")
        self.Сектор_C_опоки.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Сектор_C_опоки)

        self.Сектор_D_опоки = QLineEdit(self)
        self.Сектор_D_опоки.setPlaceholderText("Сектор D опоки")
        self.Сектор_D_опоки.setStyleSheet("padding: 10px; margin-bottom: 20px;")
        layout.addWidget(self.Сектор_D_опоки)

        # Новые поля для времени и температуры
        self.Плавка_время_прогрева_ковша = QLineEdit(self)
        self.Плавка_время_прогрева_ковша.setPlaceholderText("Время прогрева ковша (ЧЧ:ММ)")
        self.Плавка_время_прогрева_ковша.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        self.Плавка_время_прогрева_ковша.setInputMask("99:99")
        layout.addWidget(QLabel("Время прогрева ковша (ЧЧ:ММ):"))
        layout.addWidget(self.Плавка_время_прогрева_ковша)        
        
        self.Плавка_время_перемещения = QLineEdit(self)
        self.Плавка_время_перемещения.setPlaceholderText("Время перемещения (ММ:СС):")
        self.Плавка_время_перемещения.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        self.Плавка_время_перемещения.setInputMask("99:99")
        layout.addWidget(QLabel("Время перемещения (ММ:СС):"))
        layout.addWidget(self.Плавка_время_перемещения)

        self.Плавка_время_заливки = QLineEdit(self)
        self.Плавка_время_заливки.setPlaceholderText("Время заливки (ЧЧ:ММ)")
        self.Плавка_время_заливки.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        self.Плавка_время_заливки.setInputMask("99:99")
        layout.addWidget(QLabel("Время заливки (ЧЧ:ММ):"))
        layout.addWidget(self.Плавка_время_заливки)

        self.Плавка_температура_заливки = QLineEdit(self)
        self.Плавка_температура_заливки.setPlaceholderText("Температура заливки")
        self.Плавка_температура_заливки.setStyleSheet("padding: 10px; margin-bottom: 10px;")
        layout.addWidget(self.Плавка_температура_заливки)


        self.save_button = QPushButton("Сохранить", self)
        self.save_button.setObjectName("saveButton")
        self.save_button.clicked.connect(self.save_data)
        layout.addWidget(self.save_button)

        # Добавляем кнопку поиска
        self.search_button = QPushButton("Поиск и редактирование", self)
        self.search_button.setObjectName("searchButton")
        self.search_button.clicked.connect(self.show_search_dialog)
        layout.addWidget(self.search_button)

        # Добавляем обработчик изменения даты для автогенерации номера плавки
        self.Плавка_дата.dateChanged.connect(self.generate_plavka_number)
        
        # Генерируем начальный номер плавки
        self.generate_plavka_number()

        # Установка виджета прокрутки
        scroll_area.setWidget(scroll_content)
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll_area)

    def generate_plavka_number(self):
        try:
            current_month = self.Плавка_дата.date().month()
            
            if os.path.exists('plavka.xlsx'):
                df = pd.read_excel('plavka.xlsx')
                if not df.empty:
                    # Конвертируем даты в datetime
                    df['Плавка_дата'] = pd.to_datetime(df['Плавка_дата'], format='%d.%m.%Y')
                    
                    # Фильтруем записи текущего месяца
                    current_month_records = df[df['Плавка_дата'].dt.month == current_month]
                    
                    if not current_month_records.empty:
                        # Ищем последний номер для текущего месяца
                        last_numbers = []
                        for num in current_month_records['Номер_плавки']:
                            try:
                                if isinstance(num, str) and '-' in num:
                                    month, number = num.split('-')
                                    if month == str(current_month):
                                        last_numbers.append(int(number))
                            except (ValueError, TypeError):
                                continue
                        
                        next_number = max(last_numbers) + 1 if last_numbers else 1
                    else:
                        next_number = 1
                else:
                    next_number = 1
            else:
                next_number = 1
            
            # Форматируем номер плавки: месяц-номер(с ведущими нулями)
            new_plavka_number = f"{current_month}-{str(next_number).zfill(3)}"
            self.Номер_плавки.setText(new_plavka_number)
            
            # Обновляем учетный номер после генерации номера плавки
            self.update_uchet_number()
            
        except Exception as e:
            logging.error(f"Ошибка при генерации номера плавки: {str(e)}")
            self.Номер_плавки.setText("")

    def update_uchet_number(self):
        """Обновляет учетный номер на основе номера плавки"""
        try:
            plavka_number = self.Номер_плавки.text()
            if plavka_number:
                year = str(self.Плавка_дата.date().year())[-2:]  # Последние 2 цифры года
                uchet_number = f"{plavka_number}/{year}"
                return uchet_number
        except Exception as e:
            logging.error(f"Ошибка при обновлении учетного номера: {str(e)}")
        return None

    def generate_id(self, Плавка_дата, Номер_плавки):
        year = Плавка_дата.year()
        
        # Удаляем все символы, кроме цифр, и заменяем '-' на '.'
        номер_плавки = re.sub(r'[^0-9.-]', '', Номер_плавки).replace('-', '.')
        
        if номер_плавки:  # Проверяем, что номер плавки не пустой
            return f"{year}{номер_плавки}"
        else:
            QMessageBox.warning(self, "Ошибка", "Номер плавки должен содержать хотя бы одну цифру.")
        
        return None
    
    def generate_учетный_номер(self, Плавка_дата, Номер_плавки):
        # Получаем последние две цифры года
        last_two_digits_year = str(Плавка_дата.year())[-2:]
        
        # Проверяем, что номер плавки не пустой
        if Номер_плавки:  
            return f"{Номер_плавки}/{last_two_digits_year}"
        else:
            QMessageBox.warning(self, "Ошибка")
        
        return None    

    def validate_time(self, time_str):
        """Проверка корректности ввода времени в формате ЧЧ:ММ"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return True
        except ValueError:
            return False
        return False

    def check_duplicate_id(self, id_number):
        """Проверка существования ID в plavka.xlsx"""
        try:
            if not os.path.exists('plavka.xlsx'):
                return False
            
            wb = load_workbook('plavka.xlsx', read_only=True)
            ws = wb.active
            
            # Преобразуем проверяемый ID в строку для сравнения
            id_to_check = str(id_number).strip()
            
            # Проверяем первый столбец (ID)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).strip() == id_to_check:
                    wb.close()
                    return True
            
            wb.close()
            return False
        except Exception as e:
            logging.error(f"Ошибка при проверке дубликата ID: {str(e)}")
            return False

    def validate_fields(self):
        # Проверка обязательных полей
        if not self.Номер_плавки.text().strip():
            QMessageBox.warning(self, "Ошибка", "Номер плавки обязателен")
            return False
        
        # Проверка температуры заливки
        try:
            temp = float(self.Плавка_температура_заливки.text())
            if not (500 <= temp <= 2000):  # примерный диапазон
                QMessageBox.warning(self, "Ошибка", "Недопустимая температура заливки")
                return False
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Температура должна быть числом")
            return False
        
        return True

    def format_temperature(self, temp_str):
        """Форматирование температуры в нужный формат"""
        try:
            temp = float(temp_str)
            return f"{temp:.1f}°C"
        except ValueError:
            return temp_str

    def save_data(self):
        try:
            logging.info(f"Начало сохранения данных плавки {self.Номер_плавки.text()}")
            # Получаем ID из поля ввода
            id_number = self.generate_id(self.Плавка_дата.date(), self.Номер_плавки.text())
            
            # Проверяем, не пустой ли ID
            if not id_number:
                QMessageBox.warning(self, "Ошибка", "Введите ID плавки!")
                return
            
            # Проверяем на дубликат
            if self.check_duplicate_id(id_number):
                QMessageBox.warning(self, "Ошибка", 
                    f"Плавка с ID {id_number} уже существует в базе данных!")
                return
            
            # Если проверки пройдены, продолжаем сохранение
            Плавка_дата = self.Плавка_дата.date()
            formatted_date = Плавка_дата.toString("dd.MM.yyyy")
            Номер_плавки = self.Номер_плавки.text()

            Учетный_номер = self.update_uchet_number()
            if Учетный_номер is None:
                return

            Номер_кластера = self.Номер_кластера.text()
            Старший_смены_плавки = self.Старший_смены_плавки.currentText()  # Изменено на currentText()
            Первый_участник_смены_плавки = self.Первый_участник_смены_плавки.currentText()  # Изменено на currentText()
            Второй_участник_смены_плавки = self.Второй_участник_смены_плавки.currentText()  # Изменено на currentText()
            Третий_участник_смены_плавки = self.Третий_участник_смены_плавки.currentText()  # Изменено на currentText()
            Четвертый_участник_смены_плавки = self.Четвертый_участник_смены_плавки.currentText()  # Изменено на currentText()
            Наименование_отливки = self.Наименование_отливки.currentText()  # Изменено на currentText()
            Тип_эксперемента = self.Тип_эксперемента.currentText()  # Изменено на currentText()
            Сектор_A_опоки = self.Сектор_A_опоки.text()
            Сектор_B_опоки = self.Сектор_B_опоки.text()
            Сектор_C_опоки = self.Сектор_C_опоки.text()
            Сектор_D_опоки = self.Сектор_D_опоки.text()

            Плавка_время_прогрева_ковша = self.Плавка_время_прогрева_ковша.text()
            Плавка_время_перемещения = self.Плавка_время_перемещения.text()
            Плавка_время_заливки = self.Плавка_время_заливки.text()

            # Валидация времени
            if not (self.validate_time(Плавка_время_заливки) and self.validate_time(Плавка_время_прогрева_ковша) and self.validate_time(Плавка_время_перемещения)):
                QMessageBox.warning(self, "Ошибка", "Некорректный ввод времени. Используйте формат ЧЧ:ММ.")
                return

            Плавка_температура_заливки = self.Плавка_температура_заливки.text()

            save_to_excel(id_number, Учетный_номер, formatted_date, Номер_плавки, Номер_кластера,
                           Старший_смены_плавки, Первый_участник_смены_плавки,
                           Второй_участник_смены_плавки, Третий_участник_смены_плавки,
                           Четвертый_участник_смены_плавки, Наименование_отливки,
                           Тип_эксперемента, Сектор_A_опоки, Сектор_B_опоки,
                           Сектор_C_опоки, Сектор_D_опоки, Плавка_время_прогрева_ковша,
                           Плавка_время_перемещения, Плавка_время_заливки,
                           Плавка_температура_заливки)

            QMessageBox.information(self, "Успех", "Данные сохранены в Excel!")

            # Очистка полей ввода
            self.clear_fields()
            logging.info("Данные успешно сохранены")
            
            # После успешного сохранения генерируем новый номер
            self.generate_plavka_number()
            
        except Exception as e:
            logging.error(f"Ошибка при сохранении данных: {str(e)}")
            QMessageBox.critical(self, "Ошибка", str(e))

    def clear_fields(self):
        self.Плавка_дата.setDate(QDate.currentDate().addDays(-1))
        self.Номер_плавки.clear()
        self.Номер_кластера.clear()
        self.Старший_смены_плавки.setCurrentIndex(-1)  # Сброс выбора
        self.Первый_участник_смены_плавки.setCurrentIndex(-1)  # Сброс выбора
        self.Второй_участник_смены_плавки.setCurrentIndex(-1)  # Сброс выбора
        self.Третий_участник_смены_плавки.setCurrentIndex(-1)  # Сброс выбора
        self.Четвертый_участник_смены_плавки.setCurrentIndex(-1)  # Сброс выбора
        self.Наименование_отливки.setCurrentIndex(-1)  # Сброс выбора
        self.Тип_эксперемента.setCurrentIndex(-1)  # Сброс выбора
        self.Сектор_A_опоки.clear()
        self.Сектор_B_опоки.clear()
        self.Сектор_C_опоки.clear()
        self.Сектор_D_опоки.clear()
        self.Плавка_время_прогрева_ковша.clear()
        self.Плавка_время_перемещения.clear()
        self.Плавка_время_заливки.clear()
        self.Плавка_температура_заливки.clear()

    def show_search_dialog(self):
        dialog = SearchDialog(self)
        dialog.exec_()

class StatisticsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Создаем таблицу для отображения данных
        self.data_table = QTableWidget()
        self.data_table.setColumnCount(2)
        self.data_table.setHorizontalHeaderLabels(["Дата", "Температура"])
        layout.addWidget(self.data_table)
        
        # Кнопки для разных типов отображения
        buttons_layout = QHBoxLayout()
        
        temp_button = QPushButton("Температуры")
        temp_button.clicked.connect(lambda: self.show_data('temperature'))
        
        casting_button = QPushButton("Статистика отливок")
        casting_button.clicked.connect(lambda: self.show_data('castings'))
        
        time_button = QPushButton("Временной анализ")
        time_button.clicked.connect(lambda: self.show_data('time'))
        
        buttons_layout.addWidget(temp_button)
        buttons_layout.addWidget(casting_button)
        buttons_layout.addWidget(time_button)
        
        layout.addLayout(buttons_layout)
    
    def show_data(self, data_type):
        self.data_table.setRowCount(0)
        
        try:
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            if data_type == 'temperature':
                self._show_temperature(ws, headers)
            elif data_type == 'castings':
                self._show_castings(ws, headers)
            elif data_type == 'time':
                self._show_time_analysis(ws, headers)
                
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при отображении данных: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при отображении данных: {str(e)}")
    
    def _show_temperature(self, ws, headers):
        self.data_table.setHorizontalHeaderLabels(["Дата", "Температура"])
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            try:
                temp = float(data['Плавка_температура_заливки'])
                date = data['Плавка_дата']
                
                row_position = self.data_table.rowCount()
                self.data_table.insertRow(row_position)
                
                self.data_table.setItem(row_position, 0, QTableWidgetItem(date))
                self.data_table.setItem(row_position, 1, QTableWidgetItem(f"{temp}°C"))
                
            except (ValueError, TypeError):
                continue
        
        self.data_table.resizeColumnsToContents()

class SearchDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Поиск записей")
        self.setMinimumSize(1000, 700)
        
        # Добавляем тень для окна
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(0)
        shadow.setColor(QColor(0, 0, 0, 60))
        self.setGraphicsEffect(shadow)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Добавляем фильтры
        filter_group = QGroupBox("Фильтры")
        filter_layout = QGridLayout()
        
        # Фильтр по дате
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        
        filter_layout.addWidget(QLabel("Дата с:"), 0, 0)
        filter_layout.addWidget(self.date_from, 0, 1)
        filter_layout.addWidget(QLabel("по:"), 0, 2)
        filter_layout.addWidget(self.date_to, 0, 3)
        
        # Фильтр по типу отливки
        self.filter_casting = QComboBox()
        self.filter_casting.addItems(["Все"] + [
            "Вороток", "Ригель", "Ригель optima", "Блок-картер",
            "Накладка резьб", "Блок цилиндров", "Диагональ optima"
        ])
        filter_layout.addWidget(QLabel("Тип отливки:"), 1, 0)
        filter_layout.addWidget(self.filter_casting, 1, 1)
        
        # Фильтр по температуре
        self.temp_from = QLineEdit()
        self.temp_to = QLineEdit()
        filter_layout.addWidget(QLabel("Температура от:"), 2, 0)
        filter_layout.addWidget(self.temp_from, 2, 1)
        filter_layout.addWidget(QLabel("до:"), 2, 2)
        filter_layout.addWidget(self.temp_to, 2, 3)
        
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        # Добавляем вкладки для результатов и статистики
        self.tab_widget = QTabWidget()
        
        # Вкладка результатов поиска
        search_tab = QWidget()
        search_layout = QVBoxLayout(search_tab)
        
        # Существующие виджеты поиска
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите текст для поиска...")
        search_layout.addWidget(self.search_input)
        
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(len(SEARCH_FIELDS))
        self.results_table.setHorizontalHeaderLabels(SEARCH_FIELDS)
        search_layout.addWidget(self.results_table)
        
        self.tab_widget.addTab(search_tab, "Результаты поиска")
        
        # Вкладка статистики
        stats_tab = QWidget()
        stats_layout = QVBoxLayout(stats_tab)
        
        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        stats_layout.addWidget(self.stats_text)
        
        self.tab_widget.addTab(stats_tab, "Статистика")
        
        # Добавляем вкладку визуализации
        viz_tab = StatisticsWidget()
        self.tab_widget.addTab(viz_tab, "Визуализация")
        
        layout.addWidget(self.tab_widget)
        
        # Кнопки
        button_layout = QHBoxLayout()
        self.search_button = QPushButton("Поиск")
        self.edit_button = QPushButton("Редактировать")
        self.export_button = QPushButton("Экспорт")
        self.stats_button = QPushButton("Обновить статистику")
        self.backup_button = QPushButton("Создать резервную копию")
        
        button_layout.addWidget(self.search_button)
        button_layout.addWidget(self.edit_button)
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.stats_button)
        button_layout.addWidget(self.backup_button)
        layout.addLayout(button_layout)
        
        # Подключаем обработчики
        self.search_button.clicked.connect(self.search_records)
        self.edit_button.clicked.connect(self.edit_selected)
        self.export_button.clicked.connect(self.export_results)
        self.stats_button.clicked.connect(self.update_statistics)
        self.backup_button.clicked.connect(self.create_backup)

    def apply_filters(self, row, headers):
        """Применяет фильтры к записи"""
        try:
            data = dict(zip(headers, row))
            
            # Фильтр по дате
            record_date = QDate.fromString(data['Плавка_дата'], "dd.MM.yyyy")
            if not (self.date_from.date() <= record_date <= self.date_to.date()):
                return False
            
            # Фильтр по типу отливки
            if self.filter_casting.currentText() != "Все" and \
               data['Наименование_отливки'] != self.filter_casting.currentText():
                return False
            
            # Фильтр по температуре
            if self.temp_from.text() and self.temp_to.text():
                try:
                    temp = float(data['Плавка_температура_заливки'])
                    temp_from = float(self.temp_from.text())
                    temp_to = float(self.temp_to.text())
                    if not (temp_from <= temp <= temp_to):
                        return False
                except ValueError:
                    pass
            
            return True
        except Exception as e:
            logging.error(f"Ошибка при применении фильтров: {str(e)}")
            return False

    def update_statistics(self):
        """Обновляет статистику по данным"""
        try:
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            stats = {
                'total_records': 0,
                'avg_temp': [],
                'casting_types': {},
                'participants': set(),
                'min_temp': float('inf'),
                'max_temp': float('-inf')
            }
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not self.apply_filters(row, headers):
                    continue
                    
                data = dict(zip(headers, row))
                stats['total_records'] += 1
                
                # Температура
                try:
                    temp = float(data['Плавка_температура_заливки'])
                    stats['avg_temp'].append(temp)
                    stats['min_temp'] = min(stats['min_temp'], temp)
                    stats['max_temp'] = max(stats['max_temp'], temp)
                except (ValueError, TypeError):
                    pass
                
                # Типы отливок
                casting = data['Наименование_отливки']
                stats['casting_types'][casting] = stats['casting_types'].get(casting, 0) + 1
                
                # Участники
                stats['participants'].add(data['Старший_смены_плавки'])
                stats['participants'].add(data['Первый_участник_смены_плавки'])
                stats['participants'].add(data['Второй_участник_смены_плавки'])
                stats['participants'].add(data['Третий_участник_смены_плавки'])
                stats['participants'].add(data['Четвертый_участник_смены_плавки'])
            
            # Формируем отчет
            report = [
                "=== Общая статистика ===",
                f"Всего записей: {stats['total_records']}",
                f"Количество участников: {len(stats['participants'])}",
                "",
                "=== Температура заливки ===",
                f"Средняя: {sum(stats['avg_temp'])/len(stats['avg_temp']):.1f}°C" if stats['avg_temp'] else "Нет данных",
                f"Минимальная: {stats['min_temp']}°C" if stats['min_temp'] != float('inf') else "Нет данных",
                f"Максимальная: {stats['max_temp']}°C" if stats['max_temp'] != float('-inf') else "Нет данных",
                "",
                "=== Распределение по типам отливок ===",
            ]
            
            for casting, count in sorted(stats['casting_types'].items()):
                report.append(f"{casting}: {count} ({count/stats['total_records']*100:.1f}%)")
            
            self.stats_text.setText("\n".join(report))
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при обновлении статистики: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обновлении статистики: {str(e)}")

    def search_records(self):
        search_text = self.search_input.text().lower()
        try:
            wb = load_workbook(EXCEL_FILENAME, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            self.results_table.setRowCount(0)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Проверяем фильтры
                if not self.apply_filters(row, headers):
                    continue
                    
                # Ищем совпадения
                for cell in row:
                    if cell and str(cell).lower().find(search_text) != -1:
                        row_position = self.results_table.rowCount()
                        self.results_table.insertRow(row_position)
                        
                        for col, field in enumerate(SEARCH_FIELDS):
                            field_index = headers.index(field)
                            self.results_table.setItem(
                                row_position, col, 
                                QTableWidgetItem(str(row[field_index]))
                            )
                        break
            
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при поиске: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске: {str(e)}")

    def edit_selected(self):
        current_row = self.results_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите запись для редактирования")
            return
            
        # Получаем ID выбранной записи
        record_id = self.results_table.item(current_row, 0).text()
        
        # Создаем диалог редактирования
        edit_dialog = EditRecordDialog(record_id, self)
        if edit_dialog.exec_() == QDialog.Accepted:
            # Обновляем таблицу после редактирования
            self.search_records()

    def export_results(self):
        try:
            format_str = "Excel files (*.xlsx);;CSV files (*.csv);;PDF files (*.pdf)"
            file_name, selected_format = QFileDialog.getSaveFileName(
                self, "Экспорт данных", "", format_str
            )
            
            if file_name:
                # Создаем DataFrame из данных таблицы
                data = []
                for row in range(self.results_table.rowCount()):
                    row_data = []
                    for col in range(self.results_table.columnCount()):
                        item = self.results_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)
                
                df = pd.DataFrame(data, columns=SEARCH_FIELDS)
                
                if "xlsx" in selected_format:
                    df.to_excel(file_name, index=False)
                elif "csv" in selected_format:
                    df.to_csv(file_name, index=False)
                elif "pdf" in selected_format:
                    # Для PDF потребуется дополнительная настройка
                    df.to_html(file_name.replace('.pdf', '.html'))
                    # Конвертация HTML в PDF (требуется дополнительная библиотека)
                
                QMessageBox.information(self, "Успех", "Данные успешно экспортированы")
                
        except Exception as e:
            logging.error(f"Ошибка при экспорте: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте: {str(e)}")

    def create_backup(self):
        try:
            # Создаем директорию для резервных копий если её нет
            if not os.path.exists(BACKUP_DIR):
                os.makedirs(BACKUP_DIR)
            
            # Формируем имя файла резервной копии
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(BACKUP_DIR, f'plavka_backup_{timestamp}.xlsx')
            
            # Копируем файл
            shutil.copy2(EXCEL_FILENAME, backup_file)
            
            # Удаляем старые резервные копии если их больше MAX_BACKUPS
            backups = sorted([os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR)])
            while len(backups) > MAX_BACKUPS:
                os.remove(backups[0])
                backups.pop(0)
            
            QMessageBox.information(self, "Успех", 
                f"Резервная копия создана:\n{backup_file}")
            
        except Exception as e:
            logging.error(f"Ошибка при создании резервной копии: {str(e)}")
            QMessageBox.critical(self, "Ошибка", 
                f"Ошибка при создании резервной копии: {str(e)}")

class EditRecordDialog(QDialog):
    def __init__(self, record_id, parent=None):
        super().__init__(parent)
        self.record_id = record_id
        self.setWindowTitle(f"Редактирование записи {record_id}")
        self.setup_ui()
        self.load_record_data()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Создаем область прокрутки
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        content_layout = QVBoxLayout(scroll_content)
        
        # Список участников
        participants = [
            "Белков", "Карасев", "Ермаков", "Рабинович",
            "Валиулин", "Волков", "Семенов", "Левин",
            "Исмаилов", "Беляев", "Политов", "Кокшин",
            "Терентьев"
        ]
        participants.sort()
        
        # Список наименований отливок
        naimenovanie_otlivok = [
            "Вороток", "Ригель", "Ригель optima", "Блок-картер",
            "Накладка резьб", "Блок цилиндров", "Диагональ optima"
        ]
        naimenovanie_otlivok.sort()
        
        # Список типов эксперементов
        types = ["Бумага", "Волокно"]
        types.sort()
        
        # Создаем поля ввода
        self.Плавка_дата = QDateEdit(self)
        self.Плавка_дата.setDisplayFormat("dd.MM.yyyy")
        self.Плавка_дата.setCalendarPopup(True)
        content_layout.addWidget(QLabel("Дата плавки:"))
        content_layout.addWidget(self.Плавка_дата)

        self.Номер_плавки = QLineEdit(self)
        content_layout.addWidget(QLabel("Номер плавки:"))
        content_layout.addWidget(self.Номер_плавки)

        self.Номер_кластера = QLineEdit(self)
        content_layout.addWidget(QLabel("Номер кластера:"))
        content_layout.addWidget(self.Номер_кластера)

        # Комбобоксы для участников
        self.Старший_смены_плавки = QComboBox(self)
        self.Старший_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Старший смены:"))
        content_layout.addWidget(self.Старший_смены_плавки)

        self.Первый_участник_смены_плавки = QComboBox(self)
        self.Первый_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Первый участник:"))
        content_layout.addWidget(self.Первый_участник_смены_плавки)

        self.Второй_участник_смены_плавки = QComboBox(self)
        self.Второй_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Второй участник:"))
        content_layout.addWidget(self.Второй_участник_смены_плавки)

        self.Третий_участник_смены_плавки = QComboBox(self)
        self.Третий_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Третий участник:"))
        content_layout.addWidget(self.Третий_участник_смены_плавки)

        self.Четвертый_участник_смены_плавки = QComboBox(self)
        self.Четвертый_участник_смены_плавки.addItems(participants)
        content_layout.addWidget(QLabel("Четвертый участник:"))
        content_layout.addWidget(self.Четвертый_участник_смены_плавки)

        self.Наименование_отливки = QComboBox(self)
        self.Наименование_отливки.addItems(naimenovanie_otlivok)
        content_layout.addWidget(QLabel("Наименование отливки:"))
        content_layout.addWidget(self.Наименование_отливки)

        self.Тип_эксперемента = QComboBox(self)
        self.Тип_эксперемента.addItems(types)
        content_layout.addWidget(QLabel("Тип эксперимента:"))
        content_layout.addWidget(self.Тип_эксперемента)

        # Поля для секторов опоки
        self.Сектор_A_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор A опоки:"))
        content_layout.addWidget(self.Сектор_A_опоки)

        self.Сектор_B_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор B опоки:"))
        content_layout.addWidget(self.Сектор_B_опоки)

        self.Сектор_C_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор C опоки:"))
        content_layout.addWidget(self.Сектор_C_опоки)

        self.Сектор_D_опоки = QLineEdit(self)
        content_layout.addWidget(QLabel("Сектор D опоки:"))
        content_layout.addWidget(self.Сектор_D_опоки)

        # Поля времени
        self.Плавка_время_прогрева_ковша = QLineEdit(self)
        self.Плавка_время_прогрева_ковша.setInputMask("99:99")
        content_layout.addWidget(QLabel("Время прогрева ковша (ЧЧ:ММ):"))
        content_layout.addWidget(self.Плавка_время_прогрева_ковша)

        self.Плавка_температура_заливки = QLineEdit(self)
        content_layout.addWidget(QLabel("Температура заливки:"))
        content_layout.addWidget(self.Плавка_температура_заливки)

        # Кнопки
        button_layout = QHBoxLayout()
        save_button = QPushButton("Сохранить изменения")
        cancel_button = QPushButton("Отмена")
        
        save_button.clicked.connect(self.save_changes)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        
        # Устанавливаем виджеты
        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)
        layout.addLayout(button_layout)

    def load_record_data(self):
        try:
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == self.record_id:
                    # Заполняем поля данными
                    self.fill_fields(row, headers)
                    break
                    
            wb.close()
            
        except Exception as e:
            logging.error(f"Ошибка при загрузке записи: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке записи: {str(e)}")

    def fill_fields(self, row, headers):
        """Заполняет поля формы данными из записи"""
        try:
            # Создаем словарь с данными
            data = dict(zip(headers, row))
            
            # Заполняем поля
            self.Плавка_дата.setDate(QDate.fromString(data['Плавка_дата'], "dd.MM.yyyy"))
            self.Номер_плавки.setText(str(data['Номер_плавки']))
            self.Номер_кластера.setText(str(data['Номер_кластера']))
            
            # Устанавливаем значения комбобоксов
            self.Старший_смены_плавки.setCurrentText(str(data['Старший_смены_плавки']))
            self.Первый_участник_смены_плавки.setCurrentText(str(data['Первый_участник_смены_плавки']))
            self.Второй_участник_смены_плавки.setCurrentText(str(data['Второй_участник_смены_плавки']))
            self.Третий_участник_смены_плавки.setCurrentText(str(data['Третий_участник_смены_плавки']))
            self.Четвертый_участник_смены_плавки.setCurrentText(str(data['Четвертый_участник_смены_плавки']))
            
            self.Наименование_отливки.setCurrentText(str(data['Наименование_отливки']))
            self.Тип_эксперемента.setCurrentText(str(data['Тип_эксперемента']))
            
            # Заполняем секторы опоки
            self.Сектор_A_опоки.setText(str(data['Сектор_A_опоки']))
            self.Сектор_B_опоки.setText(str(data['Сектор_B_опоки']))
            self.Сектор_C_опоки.setText(str(data['Сектор_C_опоки']))
            self.Сектор_D_опоки.setText(str(data['Сектор_D_опоки']))
            
            # Заполняем время и температуру
            self.Плавка_время_прогрева_ковша.setText(str(data['Плавка_время_прогрева_ковша']))
            self.Плавка_температура_заливки.setText(str(data['Плавка_температура_заливки']))
            
        except Exception as e:
            logging.error(f"Ошибка при заполнении полей: {str(e)}")
            raise

    def save_changes(self):
        """Сохраняет изменения в Excel файл"""
        try:
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active
            
            # Находим строку с нужным ID
            row_index = None
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[0].value) == self.record_id:
                    row_index = idx
                    break
            
            if row_index:
                # Обновляем данные в строке
                ws.cell(row=row_index, column=3).value = self.Плавка_дата.date().toString("dd.MM.yyyy")
                ws.cell(row=row_index, column=4).value = self.Номер_плавки.text()
                ws.cell(row=row_index, column=5).value = self.Номер_кластера.text()
                ws.cell(row=row_index, column=6).value = self.Старший_смены_плавки.currentText()
                ws.cell(row=row_index, column=7).value = self.Первый_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=8).value = self.Второй_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=9).value = self.Третий_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=10).value = self.Четвертый_участник_смены_плавки.currentText()
                ws.cell(row=row_index, column=11).value = self.Наименование_отливки.currentText()
                ws.cell(row=row_index, column=12).value = self.Тип_эксперемента.currentText()
                ws.cell(row=row_index, column=13).value = self.Сектор_A_опоки.text()
                ws.cell(row=row_index, column=14).value = self.Сектор_B_опоки.text()
                ws.cell(row=row_index, column=15).value = self.Сектор_C_опоки.text()
                ws.cell(row=row_index, column=16).value = self.Сектор_D_опоки.text()
                ws.cell(row=row_index, column=17).value = self.Плавка_время_прогрева_ковша.text()
                ws.cell(row=row_index, column=18).value = self.Плавка_время_перемещения.text()
                ws.cell(row=row_index, column=19).value = self.Плавка_время_заливки.text()
                ws.cell(row=row_index, column=20).value = self.Плавка_температура_заливки.text()
                
                wb.save(EXCEL_FILENAME)
                QMessageBox.information(self, "Успех", "Изменения сохранены")
                self.accept()
            
        except Exception as e:
            logging.error(f"Ошибка при сохранении изменений: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении изменений: {str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
