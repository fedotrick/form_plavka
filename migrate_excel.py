from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

def convert_date_format(date_str):
    """Конвертирует дату в формат dd.MM.yyyy"""
    try:
        # Пробуем разные форматы даты
        for fmt in ["%Y-%m-%d", "%d.%m.%Y"]:
            try:
                date_obj = datetime.strptime(str(date_str), fmt)
                return date_obj.strftime("%d.%m.%Y")
            except ValueError:
                continue
        return date_str
    except:
        return date_str

def format_time(time_str):
    """Форматирует время в формат HH:MM"""
    if not time_str:
        return ""
        
    try:
        time_str = str(time_str).strip()
        if ':' not in time_str:
            return time_str
            
        hours, minutes = time_str.split(':')
        hours = int(hours)
        minutes = int(minutes)
        
        if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
            return time_str
            
        return f"{hours:02d}:{minutes:02d}"
    except:
        return time_str

def migrate_excel():
    old_file = 'plavka.xlsx'
    temp_file = 'plavka_new.xlsx'
    
    if not os.path.exists(old_file):
        print("Файл plavka.xlsx не найден")
        return
    
    # Загружаем старый файл
    old_wb = load_workbook(old_file)
    old_sheet = old_wb.active
    
    # Создаем новый файл
    new_wb = Workbook()
    new_sheet = new_wb.active
    
    # Получаем все строки из старого файла
    rows = list(old_sheet.rows)
    old_headers = [cell.value for cell in rows[0]]
    
    # Создаем новые заголовки
    new_headers = []
    for i, header in enumerate(old_headers):
        if i == 0:
            new_headers.append("id_plavka")  # Меняем ID на id_plavka
        elif header == "Номер_плавки":
            # Добавляем Плавка_время_заливки перед Номер_плавки
            new_headers.extend(["Плавка_время_заливки", header])
        else:
            new_headers.append(header)
    
    # Добавляем id в конец
    if "id" not in new_headers:
        new_headers.append("id")
    
    # Записываем заголовки
    new_sheet.append(new_headers)
    
    # Список колонок с временем
    time_columns = [
        'Плавка_время_заливки',
        'Плавка_время_прогрева_ковша_A',
        'Плавка_время_перемещения_A',
        'Плавка_время_заливки_A',
        'Плавка_время_прогрева_ковша_B',
        'Плавка_время_перемещения_B',
        'Плавка_время_заливки_B',
        'Плавка_время_прогрева_ковша_C',
        'Плавка_время_перемещения_C',
        'Плавка_время_заливки_C',
        'Плавка_время_прогрева_ковша_D',
        'Плавка_время_перемещения_D',
        'Плавка_время_заливки_D'
    ]
    
    # Копируем данные и добавляем новые колонки
    for i, row in enumerate(rows[1:], start=1):
        values = []
        old_values = [cell.value for cell in row]
        
        for j, value in enumerate(old_values):
            header = old_headers[j]
            
            if header == 'Плавка_дата':
                values.append(convert_date_format(value))
            elif header in time_columns:
                values.append(format_time(value))
            elif j == 0:
                values.append(value)  # id_plavka
            elif old_headers[j] == "Номер_плавки":
                values.extend(["", value])  # Пустое значение для Плавка_время_заливки
            else:
                values.append(value)
        
        # Добавляем id в конец
        values.append(i)
        
        new_sheet.append(values)
    
    # Сохраняем новый файл
    new_wb.save(temp_file)
    
    # Закрываем файлы
    old_wb.close()
    new_wb.close()
    
    # Заменяем старый файл новым
    os.remove(old_file)
    os.rename(temp_file, old_file)
    
    print("Миграция завершена успешно")

if __name__ == "__main__":
    migrate_excel()
